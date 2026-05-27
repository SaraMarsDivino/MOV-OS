from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from django.db import transaction
from products.models import Product
from MOVOS.money import parse_clp_pesos
from datetime import datetime, date
from django.utils.dateparse import parse_date
from decimal import Decimal
import os
import re


class Command(BaseCommand):
    help = 'Reprocesa el archivo subir.xlsx (en settings.BASE_DIR) y actualiza/crea productos usando CODIGO 1 como id.'

    def add_arguments(self, parser):
        parser.add_argument('--path', help='Ruta al archivo xlsx (relativa a BASE_DIR o absoluta)', default=None)
        parser.add_argument('--dry-run', action='store_true', help='No aplica cambios, solo muestra resumen')

    def handle(self, *args, **options):
        path = options.get('path')
        dry_run = options.get('dry_run')
        base = getattr(settings, 'BASE_DIR', '.')
        if path:
            if os.path.isabs(path):
                fp = path
            else:
                fp = os.path.join(base, path)
        else:
            fp = os.path.join(base, 'subir.xlsx')

        if not os.path.exists(fp):
            self.stdout.write(self.style.ERROR(f'Archivo no encontrado: {fp}'))
            return

        self.stdout.write(f'Procesando archivo: {fp} (dry_run={dry_run})')

        def _norm(h):
            if not h:
                return ''
            h2 = re.sub(r"\s*\(.*\)\s*$", '', str(h)).strip()
            return h2.upper()

        def safe_decimal(val):
            if val is None or (isinstance(val, str) and str(val).strip() == ''):
                return Decimal('0')
            try:
                return parse_clp_pesos(val)
            except Exception:
                try:
                    return Decimal(str(val))
                except Exception:
                    return Decimal('0')

        workbook = load_workbook(fp, read_only=True, data_only=True)
        sheet = workbook.active
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header_row_values = [str(v).strip() for v in (first_row or [])]
        header_map = { _norm(header_row_values[idx]): idx for idx in range(len(header_row_values)) if header_row_values[idx] }

        minimal_headers = ['NOMBRE', 'CODIGO 1', 'PRECIO DE COMPRA', 'PRECIO DE VENTA']
        missing = [h for h in minimal_headers if h not in header_map]
        if missing:
            self.stdout.write(self.style.ERROR(f'Faltan encabezados obligatorios: {", ".join(missing)}'))
            return

        existing_map = {p.producto_id: p for p in Product.objects.filter(producto_id__isnull=False)}

        parsed_map = {}
        rows_total = 0
        skipped_rows = []
        duplicates = []
        first_occurrence = {}
        warnings = []
        errors = []

        for row_idx, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            rows_total += 1
            if not any(v for v in row_values if v is not None and str(v).strip() != ''):
                skipped_rows.append({'row': row_idx, 'reason': 'Fila vacía'})
                continue

            def get_val(header_name):
                idx = header_map.get(header_name if header_name is None else header_name.upper())
                if idx is None:
                    idx = header_map.get(re.sub(r"\s*\(.*\)\s*$", '', str(header_name)).upper())
                if idx is not None and idx < len(row_values):
                    return row_values[idx]
                return None

            try:
                producto_id_excel = get_val('CODIGO 1')
                if producto_id_excel is None or str(producto_id_excel).strip() == '':
                    skipped_rows.append({'row': row_idx, 'reason': 'CODIGO 1 vacío'})
                    continue
                producto_id_excel = str(producto_id_excel).strip()
                if producto_id_excel in first_occurrence:
                    duplicates.append({'codigo': producto_id_excel, 'first_row': first_occurrence[producto_id_excel], 'duplicate_row': row_idx})
                    warnings.append(f'Fila {row_idx}: Código duplicado en archivo ({producto_id_excel}). Se usa la última aparición para actualizar/crear.')
                else:
                    first_occurrence[producto_id_excel] = row_idx

                nombre = str(get_val('NOMBRE')).strip() if get_val('NOMBRE') is not None else ''
                descripcion = str(get_val('DESCRIPCION')).strip() if get_val('DESCRIPCION') is not None else ''
                codigo_alternativo = None
                if 'CODIGO 2' in header_map:
                    raw_codigo2 = get_val('CODIGO 2')
                    if raw_codigo2 is not None and str(raw_codigo2).strip() != '':
                        codigo_alternativo = str(raw_codigo2).strip()

                codigo_barras_excel = None
                if 'CODIGO DE BARRAS' in header_map:
                    codigo_barras_excel = str(get_val('CODIGO DE BARRAS')).strip() if get_val('CODIGO DE BARRAS') is not None else ''
                if not codigo_barras_excel and codigo_alternativo:
                    codigo_barras_excel = codigo_alternativo

                fecha_ingreso_producto = None
                fecha_raw = get_val('FECHA DE INGRESO')
                if fecha_raw is not None and str(fecha_raw).strip() != '':
                    if isinstance(fecha_raw, (datetime, date)):
                        fecha_ingreso_producto = fecha_raw.date() if isinstance(fecha_raw, datetime) else fecha_raw
                    else:
                        try:
                            fecha_ingreso_producto = parse_date(str(fecha_raw).split(' ')[0].strip())
                        except Exception:
                            warnings.append(f'Fila {row_idx}: Fecha inválida "{fecha_raw}" -> se asigna nulo.')

                precio_compra = safe_decimal(get_val('PRECIO DE COMPRA'))
                precio_venta = safe_decimal(get_val('PRECIO DE VENTA'))

                parsed_map[producto_id_excel] = {
                    'nombre': nombre,
                    'descripcion': descripcion or None,
                    'codigo_alternativo': codigo_alternativo,
                    'codigo_barras': (codigo_barras_excel or None),
                    'fecha_ingreso_producto': fecha_ingreso_producto,
                    'precio_compra': precio_compra,
                    'precio_venta': precio_venta,
                    'permitir_venta_sin_stock': True,
                }
            except Exception as e:
                errors.append(f'Fila {row_idx}: Error inesperado -> {e}')

        to_create = []
        to_update_products = []
        unchanged_codes = []
        for codigo, defaults in parsed_map.items():
            if codigo in existing_map:
                prod = existing_map[codigo]
                changed = False
                for k, v in defaults.items():
                    if isinstance(v, str) and v.strip() == '':
                        nv = None
                    else:
                        nv = v
                    if nv is not None and getattr(prod, k) != nv:
                        setattr(prod, k, nv)
                        changed = True
                if changed:
                    to_update_products.append(prod)
                else:
                    unchanged_codes.append(codigo)
            else:
                to_create.append(Product(producto_id=codigo, **defaults))

        created_count = 0
        updated_count = 0
        created_codes = []
        updated_codes = []
        unchanged_count = len(unchanged_codes)

        if dry_run:
            created_count = len(to_create)
            updated_count = len(to_update_products)
            created_codes = [p.producto_id for p in to_create]
            updated_codes = [p.producto_id for p in to_update_products]
        else:
            try:
                with transaction.atomic():
                    if to_create:
                        Product.objects.bulk_create(to_create, batch_size=500)
                        created_count = len(to_create)
                        created_codes = [p.producto_id for p in to_create]
                    if to_update_products:
                        Product.objects.bulk_update(to_update_products, [
                            'nombre','descripcion','codigo_alternativo','codigo_barras','fecha_ingreso_producto','precio_compra','precio_venta','permitir_venta_sin_stock'
                        ], batch_size=500)
                        updated_count = len(to_update_products)
                        updated_codes = [p.producto_id for p in to_update_products]
            except Exception as e:
                self.stdout.write(self.style.ERROR(f'Error de transacción: {e}'))
                return

        try:
            workbook.close()
        except Exception:
            pass

        self.stdout.write(self.style.SUCCESS(f'Filas leídas: {rows_total}'))
        self.stdout.write(self.style.SUCCESS(f'Distintos en archivo: {len(parsed_map)}'))
        self.stdout.write(self.style.SUCCESS(f'Creados: {created_count}, Actualizados: {updated_count}, Sin cambios: {unchanged_count}'))
        if created_codes:
            self.stdout.write('Ejemplo creados: ' + ', '.join(created_codes[:20]))
        if updated_codes:
            self.stdout.write('Ejemplo actualizados: ' + ', '.join(updated_codes[:20]))
        if duplicates:
            self.stdout.write('Duplicados detectados en archivo: ' + ', '.join([f"{d['codigo']} (fila {d['duplicate_row']})" for d in duplicates[:20]]))
        if warnings:
            self.stdout.write('Advertencias: ' + '; '.join(warnings[:20]))
        if errors:
            self.stdout.write(self.style.ERROR('Errores: ' + '; '.join(errors[:20])))

        # Mostrar conteos en BD y ejemplo de productos
        total = Product.objects.count()
        with_alt = Product.objects.filter(codigo_alternativo__isnull=False).exclude(codigo_alternativo='').count()
        with_bar = Product.objects.filter(codigo_barras__isnull=False).exclude(codigo_barras='').count()
        both = Product.objects.filter(codigo_alternativo__isnull=False).exclude(codigo_alternativo='').filter(codigo_barras__isnull=False).exclude(codigo_barras='').count()
        self.stdout.write(self.style.SUCCESS(f'Total en BD: {total}, con codigo_alternativo: {with_alt}, con codigo_barras: {with_bar}, con ambos: {both}'))

        sample_missing_alt = list(Product.objects.filter(codigo_alternativo__isnull=True).values_list('pk','producto_id')[:10])
        sample_missing_bar = list(Product.objects.filter(codigo_barras__isnull=True).values_list('pk','producto_id')[:10])
        self.stdout.write('Ejemplo faltantes codigo_alternativo: ' + str(sample_missing_alt))
        self.stdout.write('Ejemplo faltantes codigo_barras: ' + str(sample_missing_bar))

        # Mostrar detalle para pk=8 si existe
        try:
            p8 = Product.objects.get(pk=8)
            self.stdout.write('Producto pk=8: ' + str({'pk': p8.pk, 'producto_id': p8.producto_id, 'nombre': p8.nombre, 'codigo_alternativo': p8.codigo_alternativo, 'codigo_barras': p8.codigo_barras}))
        except Product.DoesNotExist:
            self.stdout.write('Producto pk=8 no existe en BD')
