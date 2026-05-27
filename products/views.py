from MOVOS.money import parse_clp_pesos
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger 
from django.db import transaction
from django.db.models import Q 
from .models import Product, StockSucursal, TransferenciaStock, AjusteStock
from .utils import build_product_search_q
from .forms import ProductForm
from django.contrib import messages
from django.http import HttpResponse
from django.http import JsonResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import datetime, date
from django.utils.dateparse import parse_date 
from sucursales.models import Sucursal
from django.conf import settings
from users.models import Vendedor
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_http_methods
import json
import re

from MOVOS.react_ui import react_ui_enabled, vite_dev_server_url_if_up


def _is_staff(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def _can_manage_products(user):
    """Users who can manage products: staff/superuser or explicit product permission flags."""
    if not getattr(user, 'is_authenticated', False):
        return False
    if user.is_staff or user.is_superuser:
        return True
    return bool(getattr(user, 'can_add_products', False)) or bool(getattr(user, 'can_edit_products', False))


def _wants_json(request) -> bool:
    ct = (request.content_type or '').lower()
    accept = (request.headers.get('Accept') or '').lower()
    return 'application/json' in ct or 'application/json' in accept

@user_passes_test(_can_manage_products, login_url='cashier_dashboard')
@login_required
def product_management(request):
    """
    Vista para la gestión y listado de productos, incluyendo búsqueda, paginación y ordenamiento.
    """
    # React UI: dev uses Vite dev server; prod uses built assets when enabled.
    # React UI: dev uses Vite dev server; prod uses built assets when enabled.
    # Pop the upload report so it's shown only once after an import
    upload_report = request.session.pop('upload_products_report', None)
    if react_ui_enabled():
        ctx = {}
        if getattr(settings, 'DEBUG', False):
            vite_url = vite_dev_server_url_if_up()
            if vite_url:
                ctx['vite_dev_server_url'] = vite_url
        # expose upload_report so the React app can display it while keeping React styling
        ctx['upload_report'] = upload_report
        # expose user permissions so React can enable/disable action buttons
        try:
            ctx['react_context'] = {
                'user': {
                    'is_staff': bool(getattr(request.user, 'is_staff', False)),
                    'is_superuser': bool(getattr(request.user, 'is_superuser', False)),
                    'can_add_products': bool(getattr(request.user, 'can_add_products', False)),
                    'can_edit_products': bool(getattr(request.user, 'can_edit_products', False)),
                },
            }
        except Exception:
            ctx['react_context'] = {}
        return render(request, 'products/react_product_management.html', ctx)

    query = request.GET.get('search', '')
    
    # Obtener parámetros de ordenamiento
    sort_by = request.GET.get('sort_by', 'nombre')
    order = request.GET.get('order', 'asc')

    # Filtrar productos
    products = Product.objects.filter(build_product_search_q(query))

    # Lista de campos permitidos para ordenar
    allowed_sort_fields = {
        'nombre': 'nombre',
        'descripcion': 'descripcion',
        'codigo1': 'producto_id',
        'codigo_barras': 'codigo_barras',
        'fecha_ingreso': 'fecha_ingreso_producto',
        'precio_compra': 'precio_compra',
        'precio_venta': 'precio_venta',
        'cantidad': 'cantidad',
        'stock': 'stock',
    }

    # Aplicar ordenamiento
    if sort_by in allowed_sort_fields:
        field_to_sort = allowed_sort_fields[sort_by]
        if order == 'desc':
            field_to_sort = '-' + field_to_sort
        products = products.order_by(field_to_sort)
    else:
        products = products.order_by('nombre') 

    total_products_count = products.count()
    per_page_options = [10, 25, 50, 100] 
    per_page = request.GET.get('per_page', 10)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10 
    
    if per_page not in per_page_options:
        per_page = 10 

    paginator = Paginator(products, per_page) 
    page = request.GET.get('page', 1) 
    
    try:
        products_page = paginator.page(page) 
    except PageNotAnInteger:
        products_page = paginator.page(1)
    except EmptyPage:
        products_page = paginator.page(paginator.num_pages)

    return render(request, 'products/product_management.html', {
        'products': products_page,
        'total_products_count': total_products_count,
        'upload_report': request.session.get('upload_products_report'),
        'search_query': query,
        'per_page': per_page, 
        'per_page_options': per_page_options, 
        'sort_by': sort_by, 
        'order': order,
        # Filtrado de mensajes para evitar mostrar avisos de caja aquí
        'messages': [m for m in messages.get_messages(request) if not (
            str(m).startswith('Ya existe una caja abierta') or str(m).startswith('Caja abierta en sucursal') or str(m).startswith('Ya tienes una caja abierta')
        )]
    })

@user_passes_test(_can_manage_products, login_url='cashier_dashboard')
@login_required
def create_or_edit_product(request, product_id=None):
    """
    Vista para crear un nuevo producto o editar uno existente.
    """
    product = get_object_or_404(Product, id=product_id) if product_id else None
    form = ProductForm(request.POST or None, instance=product, user=request.user)
    title = 'Editar Producto' if product_id else 'Crear Producto'

    if request.method == 'POST' and form.is_valid():
        product_instance = form.save() 
        messages.success(request, 'Los cambios se guardaron con éxito.')
        if 'save_and_list' in request.POST:
            return redirect('product_management')
        return redirect('edit_product', product_instance.id)
    # Construir resumen de stock por sucursal (si existe el producto)
    stocks = []
    if product and product.id:
        registros = StockSucursal.objects.select_related('sucursal').filter(producto=product)
        for r in registros:
            stocks.append({ 'sucursal': r.sucursal, 'cantidad': r.cantidad })
        # Fallback al stock legado si no hay registros por sucursal
        if not registros.exists() and product.sucursal_id:
            try:
                suc = Sucursal.objects.get(id=product.sucursal_id)
                stocks.append({ 'sucursal': suc, 'cantidad': product.stock })
            except Sucursal.DoesNotExist:
                pass
    return render(request, 'products/product_form.html', {
        'form': form, 
        'title': title, 
        'product': product,
        'stocks_sucursal': stocks,
        'sucursales': Sucursal.objects.all() if product and product.id else Sucursal.objects.none()
    })

@user_passes_test(_can_manage_products, login_url='cashier_dashboard')
@login_required
def delete_product(request, product_id):
    """
    Deshabilita (soft-delete) un producto para no romper ventas/reportes.
    """
    product = get_object_or_404(Product, id=product_id)
    if request.method != 'POST':
        messages.info(request, 'Usa el botón de acción en la lista para confirmar.')
        return redirect('product_management')

    if not getattr(product, 'activo', True):
        messages.info(request, 'El producto ya está deshabilitado.')
        return redirect('product_management')

    product.activo = False
    product.save(update_fields=['activo'])
    messages.success(request, 'Producto deshabilitado correctamente.')
    return redirect('product_management')


@user_passes_test(_can_manage_products, login_url='cashier_dashboard')
@login_required
@require_http_methods(["POST"])
def set_product_active(request, product_id):
    """Activa/desactiva un producto.

    - HTML form: POST active=1/0 -> redirect con mensaje.
    - JSON: POST {"active": true/false} -> JSON.
    """

    product = get_object_or_404(Product, id=product_id)

    try:
        if (request.content_type or '').lower().startswith('application/json'):
            payload = json.loads(request.body.decode('utf-8') or '{}')
            desired = payload.get('active')
        else:
            desired = request.POST.get('active')
        active = bool(desired) if isinstance(desired, bool) else str(desired).strip().lower() in ('1', 'true', 'yes', 'on')
    except Exception:
        if _wants_json(request):
            return JsonResponse({'success': False, 'error': 'Payload inválido'}, status=400)
        messages.error(request, 'Payload inválido.')
        return redirect('product_management')

    if getattr(product, 'activo', True) == active:
        msg = 'El producto ya estaba activo.' if active else 'El producto ya estaba deshabilitado.'
        if _wants_json(request):
            return JsonResponse({'success': True, 'id': product.id, 'activo': getattr(product, 'activo', True), 'message': msg})
        messages.info(request, msg)
        return redirect('product_management')

    product.activo = active
    product.save(update_fields=['activo'])

    msg = 'Producto habilitado correctamente.' if active else 'Producto deshabilitado correctamente.'
    if _wants_json(request):
        return JsonResponse({'success': True, 'id': product.id, 'activo': product.activo, 'message': msg})
    messages.success(request, msg)
    return redirect('product_management')

@user_passes_test(_is_staff, login_url='cashier_dashboard')
@login_required
def download_template(request):
    """
    Vista para descargar una plantilla Excel con los encabezados de los productos, 
    incluyendo los nuevos campos calculados.
    """
    from openpyxl import Workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_productos.xlsx"'

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Productos'

    # Nuevos encabezados que incluyen los cálculos
    headers = [
        'NOMBRE', 'DESCRIPCION (opcional)', 'CODIGO 1', 'CODIGO 2 (opcional)', 'CODIGO DE BARRAS (opcional)',
        'FECHA DE INGRESO (opcional)', 'PRECIO DE COMPRA', 'PRECIO DE VENTA',
        'PRECIO COMPRA SIN IVA', 'PRECIO VENTA SIN IVA',
        'GANANCIA NETA', 'PORCENTAJE DE GANANCIA'
    ]
    sheet.append(headers)

    header_border = Border(
        left=Side(style='thin', color='FF000000'),
        right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'),
        bottom=Side(style='thin', color='FF000000'),
    )
    header_fill = PatternFill(fill_type='solid', start_color='FFF8F8F8', end_color='FFF8F8F8')
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.border = header_border
        cell.fill = header_fill
        cell.font = header_font

    for idx, width in enumerate([30, 40, 20, 20, 25, 22, 18, 18, 24, 24, 18, 20], start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    workbook.save(response)
    return response

@user_passes_test(_is_staff, login_url='cashier_dashboard')
@login_required
def upload_products(request):
    """
    Vista para subir productos desde un archivo Excel.
    """
    if request.method == 'POST':
        dry_run = 'dry_run' in request.POST  # Permite previsualización sin escribir
        if 'file' not in request.FILES:
            messages.error(request, 'No se subió ningún archivo.')
            return redirect('upload_products')

        file = request.FILES['file']
        if not file.name.lower().endswith('.xlsx'):
            messages.error(request, 'Formato inválido. Se requiere un .xlsx')
            return redirect('upload_products')

        try:
            # Use read_only streaming mode to reduce memory usage on low-RAM instances
            workbook = load_workbook(file, read_only=True, data_only=True)
            sheet = workbook.active

            # In read_only mode, avoid random access like sheet[1]; iterate the first row instead
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            header_row_values = [str(v).strip() for v in (first_row or [])]
            # Normalize headers: remove parenthetical notes like "(opcional)" and uppercase
            def _norm(h):
                if not h:
                    return ''
                # remove trailing parenthesis notes and normalize whitespace
                h2 = re.sub(r"\s*\(.*\)\s*$", '', str(h)).strip()
                return h2.upper()

            header_map = { _norm(header_row_values[idx]): idx for idx in range(len(header_row_values)) if header_row_values[idx] }
            minimal_headers = ['NOMBRE', 'CODIGO 1', 'PRECIO DE COMPRA', 'PRECIO DE VENTA']
            missing = [h for h in minimal_headers if h not in header_map]
            if missing:
                messages.error(request, f'Faltan encabezados obligatorios: {", ".join(missing)}')
                return redirect('upload_products')

            # Pre-cargar productos existentes para minimizar queries repetidas
            existing_map = {p.producto_id: p for p in Product.objects.filter(producto_id__isnull=False)}

            # Acumuladores y métricas del archivo
            warnings = []
            errors = []
            parsed_map = {}  # mantiene la última fila leída por producto_id (el último valor gana)
            rows_total = 0
            skipped_rows = []  # filas completamente vacías o sin CODIGO 1
            duplicates = []  # lista de dicts {codigo, first_row, duplicate_row}
            first_occurrence = {}
            file_codes = []

            def safe_decimal(val):
                if val is None or (isinstance(val, str) and str(val).strip() == ''):
                    return Decimal('0')
                try:
                    return parse_clp_pesos(val)
                except Exception:
                    return Decimal('0')

            for row_idx, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                rows_total += 1
                # Salta filas completamente vacías
                if not any(v for v in row_values if v is not None and str(v).strip() != ''):
                    skipped_rows.append({'row': row_idx, 'reason': 'Fila vacía'})
                    continue

                def get_val(header_name):
                    idx = header_map.get(header_name if header_name is None else header_name.upper())
                    # also try normalized form (strip parenthesis) just in case
                    if idx is None:
                        idx = header_map.get(re.sub(r"\s*\(.*\)\s*$", '', str(header_name)).upper())
                    if idx is not None and idx < len(row_values):
                        return row_values[idx]
                    return None

                try:
                    producto_id_excel = get_val('CODIGO 1')
                    if producto_id_excel is None or str(producto_id_excel).strip() == '':
                        skipped_rows.append({'row': row_idx, 'reason': 'CODIGO 1 vacío'})
                        continue
                    producto_id_excel = str(producto_id_excel).strip()
                    file_codes.append(producto_id_excel)
                    if producto_id_excel in first_occurrence:
                        duplicates.append({'codigo': producto_id_excel, 'first_row': first_occurrence[producto_id_excel], 'duplicate_row': row_idx})
                        warnings.append(f'Fila {row_idx}: Código duplicado en archivo ({producto_id_excel}). Se usa la última aparición para actualizar/crear.')
                    else:
                        first_occurrence[producto_id_excel] = row_idx

                    nombre = str(get_val('NOMBRE')).strip() if get_val('NOMBRE') is not None else ''
                    descripcion = str(get_val('DESCRIPCION')).strip() if get_val('DESCRIPCION') is not None else ''
                    codigo_alternativo = None
                    if 'CODIGO 2' in header_map:
                        raw_codigo2 = get_val('CODIGO 2')
                        if raw_codigo2 is not None and str(raw_codigo2).strip() != '':
                            codigo_alternativo = str(raw_codigo2).strip()

                    codigo_barras_excel = None
                    if 'CODIGO DE BARRAS' in header_map:
                        codigo_barras_excel = str(get_val('CODIGO DE BARRAS')).strip() if get_val('CODIGO DE BARRAS') is not None else ''
                    if not codigo_barras_excel and codigo_alternativo:
                        # Compatibilidad hacia atrás: si solo se entregó CODIGO 2, úsalo como código de barras cuando no hay otra columna.
                        codigo_barras_excel = codigo_alternativo

                    fecha_ingreso_producto = None
                    fecha_raw = get_val('FECHA DE INGRESO')
                    if fecha_raw is not None and str(fecha_raw).strip() != '':
                        if isinstance(fecha_raw, (datetime, date)):
                            fecha_ingreso_producto = fecha_raw.date() if isinstance(fecha_raw, datetime) else fecha_raw
                        else:
                            try:
                                fecha_ingreso_producto = parse_date(str(fecha_raw).split(' ')[0].strip())
                            except Exception:
                                warnings.append(f'Fila {row_idx}: Fecha inválida "{fecha_raw}" -> se asigna nulo.')

                    precio_compra = safe_decimal(get_val('PRECIO DE COMPRA'))
                    precio_venta = safe_decimal(get_val('PRECIO DE VENTA'))

                    parsed_map[producto_id_excel] = {
                        'nombre': nombre,
                        'descripcion': descripcion or None,
                        'codigo_alternativo': codigo_alternativo,
                        'codigo_barras': (codigo_barras_excel or None),
                        'fecha_ingreso_producto': fecha_ingreso_producto,
                        'precio_compra': precio_compra,
                        'precio_venta': precio_venta,
                        'permitir_venta_sin_stock': True,
                    }
                except Exception as e:
                    errors.append(f'Fila {row_idx}: Error inesperado -> {e}')

            # Construir listas finales de creación/actualización a partir de parsed_map
            to_create = []
            to_update = []
            unchanged_codes = []
            to_update_products = []
            # Apply parsed_map: treat CODIGO 1 (producto_id) as the identifier.
            # If producto exists, update its fields with non-empty values from the file.
            for codigo, defaults in parsed_map.items():
                if codigo in existing_map:
                    prod = existing_map[codigo]
                    changed = False
                    for k, v in defaults.items():
                        # Normalize empty strings to None so we don't overwrite with blanks
                        if isinstance(v, str) and v.strip() == '':
                            nv = None
                        else:
                            nv = v
                        # Only update when the file provides a non-null value different from current
                        if nv is not None and getattr(prod, k) != nv:
                            setattr(prod, k, nv)
                            changed = True
                    if changed:
                        to_update_products.append(prod)
                    else:
                        unchanged_codes.append(codigo)
                else:
                    to_create.append(Product(producto_id=codigo, **defaults))

            created_count = 0
            updated_count = 0
            created_codes = []
            updated_codes = []
            unchanged_count = len(unchanged_codes)
            if not dry_run:
                from django.db import transaction
                try:
                    with transaction.atomic():
                        if to_create:
                            Product.objects.bulk_create(to_create, batch_size=500)
                            created_count = len(to_create)
                            created_codes = [p.producto_id for p in to_create]
                        if to_update_products:
                            Product.objects.bulk_update(to_update_products, [
                                'nombre','descripcion','codigo_alternativo','codigo_barras','fecha_ingreso_producto','precio_compra','precio_venta','permitir_venta_sin_stock'
                            ], batch_size=500)
                            updated_count = len(to_update_products)
                            updated_codes = [p.producto_id for p in to_update_products]
                except Exception as e:
                    messages.error(request, f'Error de transacción: {e}')
                    return redirect('upload_products')
            else:
                created_count = len(to_create)
                updated_count = len(to_update_products)
                created_codes = [p.producto_id for p in to_create]
                updated_codes = [p.producto_id for p in to_update_products]

            # Close workbook explicitly to free resources early
            try:
                workbook.close()
            except Exception:
                pass

            distinct_processed = len(parsed_map)
            total_processed = created_count + updated_count
            if dry_run:
                messages.info(request, f'Dry-run: {total_processed} filas procesables (Nuevos: {created_count}, Modificados: {updated_count}).')
            else:
                messages.success(request, f'Productos creados: {created_count}, actualizados: {updated_count}. Total: {total_processed}.')

            # Compactar warnings y errores para no saturar UI
            if warnings:
                messages.warning(request, f'{len(warnings)} advertencias. Ejemplo: {warnings[0]}')
            if errors:
                messages.error(request, f'{len(errors)} errores. Ejemplo: {errors[0]}')

            # Guardar detalle en sesión para revisión futura
            request.session['upload_products_report'] = {
                'rows_total': rows_total,
                'distinct_product_codes': len(parsed_map),
                'created': created_count,
                'updated': updated_count,
                'unchanged': unchanged_count,
                'created_codes': created_codes,
                'updated_codes': updated_codes,
                'unchanged_codes': unchanged_codes,
                'duplicates': duplicates,
                'skipped_rows': skipped_rows,
                'warnings': warnings[:500],  # Limitar
                'errors': errors[:500],
                'dry_run': dry_run,
                'file_codes_sample': file_codes[:200],
            }

            # Redirigir a gestión si se ejecutó realmente
            if not dry_run:
                return redirect('product_management')
            return redirect('upload_products')
        except Exception as e:
            messages.error(request, f'Error general procesando el archivo: {e}')
            return redirect('upload_products')

    return render(request, 'products/upload_products.html')

@user_passes_test(_is_staff, login_url='cashier_dashboard')
@login_required
def delete_all_products(request):
    """
    Vista para deshabilitar todos los productos.
    """
    if request.method == 'POST':
        try:
            count = Product.objects.filter(activo=True).update(activo=False)
            messages.success(request, f'¡Se deshabilitaron {count} productos exitosamente!')
            return redirect('product_management')
        except Exception as e:
            messages.error(request, f'Error al intentar eliminar productos: {str(e)}')
            return redirect('product_management')
    return render(request, 'products/delete_all_products_confirm.html')


@user_passes_test(_is_staff, login_url='cashier_dashboard')
@login_required
def bulk_delete_products(request):
    """
    Endpoint para deshabilitar varios productos a la vez desde la vista de gestión.
    Espera un POST con JSON: { "product_ids": [1,2,3] }
    Devuelve JSON con el número de deshabilitados.
    """
    if request.method == 'POST':
        try:
            # admitir tanto application/json como form data
            if request.content_type == 'application/json':
                payload = json.loads(request.body.decode('utf-8') or '{}')
                ids = payload.get('product_ids', [])
            else:
                ids = request.POST.getlist('product_ids')
            # Normalizar IDs a enteros
            ids = [int(x) for x in ids if x]
            qs = Product.objects.filter(id__in=ids)
            count = qs.filter(activo=True).update(activo=False)
            try:
                messages.success(request, f'Se deshabilitaron {count} productos correctamente.')
            except Exception:
                # En entornos de prueba RequestFactory puede no tener MessageMiddleware
                pass
            # Mantener clave 'deleted' por compatibilidad con frontend ya desplegado.
            return JsonResponse({'success': True, 'deleted': count, 'disabled': count})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

@user_passes_test(_is_staff, login_url='cashier_dashboard')
@login_required
def export_products_to_excel(request):
    """
    Vista para exportar todos los productos a un archivo Excel.
    Se utiliza el formateo definido en el modelo, incluyendo los nuevos cálculos.
    """
    from openpyxl import Workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="export_productos.xlsx"'

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Productos'

    # Encabezados actualizados
    headers = [
        'NOMBRE', 'DESCRIPCION', 'CODIGO 1', 'CODIGO DE BARRAS',
        'FECHA DE INGRESO', 'PRECIO DE COMPRA', 'PRECIO DE VENTA',
        'PRECIO COMPRA SIN IVA', 'PRECIO VENTA SIN IVA',
        'GANANCIA NETA', 'PORCENTAJE DE GANANCIA'
    ]
    sheet.append(headers)

    products = Product.objects.all().order_by('nombre')
    for product in products:
        row_data = [
            product.nombre,
            product.descripcion,
            product.producto_id,
            product.codigo_barras,
            product.fecha_ingreso_producto,
            product.formatted_precio_compra,
            product.formatted_precio_venta,
            product.formatted_precio_compra_sin_iva,
            product.formatted_precio_venta_sin_iva,
            product.formatted_ganancia_neta,
            product.porcentaje_ganancia  # O, si se tiene formateado: product.formatted_porcentaje_ganancia
        ]
        sheet.append(row_data)

    workbook.save(response)
    return response


def get_page_range(page_obj, block_size=20):
    total_pages = page_obj.paginator.num_pages
    current = page_obj.number
    start = ((current - 1) // block_size) * block_size + 1
    end = min(start + block_size - 1, total_pages)
    return range(start, end + 1)

@user_passes_test(_is_staff, login_url='cashier_dashboard')
@login_required
def bulk_assign_products(request):
    """
    Vista para asignar de forma masiva productos a una sucursal.
    """
    # Filtros
    search_query = request.GET.get('search', '')
    per_page_options = [10, 20, 30, 50, 100]
    try:
        per_page = int(request.GET.get('per_page', 10))
    except ValueError:
        per_page = 10
    if per_page not in per_page_options:
        per_page = 10

    # Queryset de productos filtrados
    products_qs = Product.objects.all()
    if search_query:
        products_qs = products_qs.filter(build_product_search_q(search_query))
    # Asegurar orden determinístico antes de paginar para evitar UnorderedObjectListWarning
    products_qs = products_qs.order_by('nombre', 'id')
    
    paginator = Paginator(products_qs, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    block_range = get_page_range(page_obj, 20)

    if request.method == "POST":
        # Si viene desde React (modal) se envía JSON y se espera JSON
        if request.content_type and request.content_type.startswith('application/json'):
            try:
                payload = json.loads(request.body.decode('utf-8') or '{}')
                sucursal_id = int(payload.get('sucursal_id') or 0)
                items = payload.get('items') or []
                if not sucursal_id or not isinstance(items, list):
                    return JsonResponse({'success': False, 'error': 'Payload inválido'}, status=400)
            except Exception as e:
                return JsonResponse({'success': False, 'error': f'Payload inválido: {e}'}, status=400)

            try:
                sucursal = Sucursal.objects.get(id=sucursal_id)
            except Sucursal.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Sucursal no encontrada'}, status=404)

            assigned = 0
            failures = []
            with transaction.atomic():
                for it in items:
                    try:
                        if not isinstance(it, dict):
                            failures.append({'item': it, 'reason': 'Item inválido'})
                            continue
                        codigo = (it.get('codigo') or it.get('producto_id') or it.get('product_code') or '').strip()
                        pid_raw = it.get('id') or it.get('product_id')
                        try:
                            pid = int(pid_raw) if pid_raw is not None else None
                        except (TypeError, ValueError):
                            pid = None

                        if it.get('cantidad') is None:
                            cantidad = 0
                        else:
                            try:
                                cantidad = int(it.get('cantidad') or 0)
                            except (TypeError, ValueError):
                                failures.append({'codigo': codigo or None, 'reason': 'Cantidad inválida'})
                                continue

                        allow_raw = it.get('permitir_venta_sin_stock')
                        allow_flag = None
                        if allow_raw is not None:
                            if isinstance(allow_raw, bool):
                                allow_flag = allow_raw
                            elif isinstance(allow_raw, str):
                                allow_flag = allow_raw.strip().lower() in ('1', 'true', 'yes', 'on')
                            else:
                                try:
                                    allow_flag = bool(int(allow_raw))
                                except (TypeError, ValueError):
                                    failures.append({'codigo': codigo or None, 'reason': 'Valor permitir_venta_sin_stock inválido'})
                                    continue

                        if not codigo and not pid:
                            failures.append({'codigo': codigo or None, 'reason': 'Código o ID requerido'})
                            continue
                        if cantidad < 0:
                            failures.append({'codigo': codigo or None, 'reason': 'Cantidad inválida'})
                            continue

                        if pid:
                            prod = Product.objects.filter(id=pid).first()
                        else:
                            prod = Product.objects.filter(Q(producto_id=codigo) | Q(codigo_barras=codigo)).first()
                        if not prod:
                            failures.append({'codigo': codigo or pid, 'reason': 'Producto no existe'})
                            continue

                        # Asegurar vínculo con sucursal y registro de stock aunque la cantidad sea 0.
                        if prod.sucursal_id != sucursal.id:
                            prod.sucursal = sucursal
                            prod.save(update_fields=['sucursal'])

                        if cantidad == 0:
                            ss, _ = StockSucursal.objects.get_or_create(
                                producto=prod,
                                sucursal=sucursal,
                                defaults={'cantidad': 0, 'permitir_venta_sin_stock': allow_flag}
                            )
                        else:
                            prod.incrementar_stock_en(sucursal, cantidad)
                            ss = StockSucursal.objects.filter(producto=prod, sucursal=sucursal).first()

                        if allow_flag is not None:
                            if ss:
                                if ss.permitir_venta_sin_stock != allow_flag:
                                    ss.permitir_venta_sin_stock = allow_flag
                                    ss.save(update_fields=['permitir_venta_sin_stock'])
                            else:
                                StockSucursal.objects.update_or_create(
                                    producto=prod,
                                    sucursal=sucursal,
                                    defaults={'cantidad': 0, 'permitir_venta_sin_stock': allow_flag}
                                )

                        assigned += 1
                    except Exception as e:
                        failures.append({'item': it, 'reason': str(e)})

            return JsonResponse({'success': True, 'assigned_count': assigned, 'failures': failures})

        # Obtener IDs de los productos marcados y la sucursal seleccionada
        product_ids = request.POST.getlist('products')
        sucursal_id = request.POST.get('sucursal')
        cantidad_str = request.POST.get('cantidad', '').strip()
        allow_mode = (request.POST.get('allow_without_stock') or '').strip().lower()
        allow_flag = None
        if allow_mode in ('allow', 'permitir', 'si', 'yes', 'true', '1'):
            allow_flag = True
        elif allow_mode in ('deny', 'no', 'false', '0'):
            allow_flag = False
        if not product_ids or not sucursal_id:
            messages.error(request, "Debe seleccionar al menos un producto y una sucursal.")
        else:
            sucursal = get_object_or_404(Sucursal, id=sucursal_id)
            products = list(Product.objects.filter(id__in=product_ids))
            if not products:
                messages.error(request, "No se encontraron productos válidos para asignar.")
            else:
                cantidad_val = None
                if cantidad_str:
                    try:
                        cantidad_val = int(cantidad_str)
                        if cantidad_val < 0:
                            raise ValueError("La cantidad no puede ser negativa.")
                    except ValueError:
                        messages.warning(request, "La cantidad ingresada no es válida. Se ignoró la actualización de stock.")
                        cantidad_val = None

                for producto in products:
                    producto.sucursal = sucursal
                    producto.save(update_fields=['sucursal'])
                    ss = None
                    if cantidad_val is not None:
                        defaults = {'cantidad': cantidad_val}
                        if allow_flag is not None:
                            defaults['permitir_venta_sin_stock'] = allow_flag
                        ss, created = StockSucursal.objects.get_or_create(
                            producto=producto,
                            sucursal=sucursal,
                            defaults=defaults
                        )
                        if not created and ss.cantidad != cantidad_val:
                            delta = cantidad_val - ss.cantidad
                            ss.cantidad = cantidad_val
                            ss.save(update_fields=['cantidad'])
                            AjusteStock.objects.create(
                                producto=producto,
                                sucursal=sucursal,
                                cantidad_delta=delta,
                                motivo='Ajuste por asignación masiva de stock',
                                usuario=request.user if request.user.is_authenticated else None
                            )
                        elif created and cantidad_val != 0:
                            AjusteStock.objects.create(
                                producto=producto,
                                sucursal=sucursal,
                                cantidad_delta=cantidad_val,
                                motivo='Asignación masiva de stock',
                                usuario=request.user if request.user.is_authenticated else None
                            )
                    if allow_flag is not None:
                        if ss is None:
                            ss, _ = StockSucursal.objects.get_or_create(
                                producto=producto,
                                sucursal=sucursal,
                                defaults={'cantidad': 0, 'permitir_venta_sin_stock': allow_flag}
                            )
                        if ss.permitir_venta_sin_stock != allow_flag:
                            ss.permitir_venta_sin_stock = allow_flag
                            ss.save(update_fields=['permitir_venta_sin_stock'])

                messages.success(request, "Productos asignados correctamente a la sucursal.")
                return redirect('product_management')

    context = {
        'page_obj': page_obj,
        'block_range': block_range,
        'search_query': search_query,
        'per_page_options': per_page_options,
        'per_page': per_page,
        'sucursales': Sucursal.objects.all(),
    }
    return render(request, 'products/bulk_assign_products.html', context)

@user_passes_test(_can_manage_products, login_url='cashier_dashboard')
@login_required
def transfer_stock(request):
    """Transferir cantidad de un producto desde una sucursal origen a una sucursal destino."""
    if react_ui_enabled():
        products_qs = Product.objects.filter(activo=True).order_by('nombre').values('id', 'nombre', 'producto_id')
        sucursales_qs = Sucursal.objects.all().order_by('nombre').values('id', 'nombre')
        ctx = {
            'react_context': {
                'products': list(products_qs),
                'sucursales': list(sucursales_qs),
            }
        }
        vite_url = vite_dev_server_url_if_up()
        if vite_url:
            ctx['vite_dev_server_url'] = vite_url
        return render(request, 'products/react_transfer_stock.html', ctx)

    # Legacy form-based path (only reached when React UI is disabled)
    productos = Product.objects.all().order_by('nombre')
    sucursales = Sucursal.objects.all().order_by('nombre')
    pre_producto = request.GET.get('producto')
    pre_origen = request.GET.get('sucursal_origen')
    context = {'productos': productos, 'sucursales': sucursales, 'pre_producto': pre_producto, 'pre_origen': pre_origen}
    if request.method == 'POST':
        try:
            producto_id = int(request.POST.get('producto_id'))
            origen_id = int(request.POST.get('sucursal_origen'))
            destino_id = int(request.POST.get('sucursal_destino'))
            cantidad = int(request.POST.get('cantidad'))
        except (TypeError, ValueError):
            messages.error(request, 'Datos inválidos en el formulario.')
            return render(request, 'products/transfer_stock.html', context)
        if origen_id == destino_id:
            messages.error(request, 'La sucursal de origen y destino no pueden ser la misma.')
            return render(request, 'products/transfer_stock.html', context)
        if cantidad <= 0:
            messages.error(request, 'La cantidad debe ser mayor a cero.')
            return render(request, 'products/transfer_stock.html', context)
        producto = get_object_or_404(Product, id=producto_id)
        suc_origen = get_object_or_404(Sucursal, id=origen_id)
        suc_destino = get_object_or_404(Sucursal, id=destino_id)
        disp = producto.stock_en(suc_origen)
        allow_without_stock = producto.permitir_venta_sin_stock_en(suc_origen)
        if disp < cantidad and not allow_without_stock:
            messages.error(request, f'Stock insuficiente en sucursal origen. Disponible: {disp}.')
            return render(request, 'products/transfer_stock.html', context)
        ss_origen, _ = StockSucursal.objects.get_or_create(producto=producto, sucursal=suc_origen, defaults={'cantidad': 0})
        ss_origen.cantidad = max(0, (ss_origen.cantidad or 0) - cantidad)
        ss_origen.save()
        ss_destino, _ = StockSucursal.objects.get_or_create(producto=producto, sucursal=suc_destino, defaults={'cantidad': 0})
        ss_destino.cantidad = (ss_destino.cantidad or 0) + cantidad
        ss_destino.save()
        TransferenciaStock.objects.create(
            producto=producto, origen=suc_origen, destino=suc_destino,
            cantidad=cantidad, usuario=request.user if request.user.is_authenticated else None
        )
        messages.success(request, f'Transferencia realizada: {cantidad} unidades de "{producto.nombre}" de {suc_origen.nombre} a {suc_destino.nombre}.')
        return redirect('transfer_stock')
    return render(request, 'products/transfer_stock.html', context)


@user_passes_test(_can_manage_products, login_url='cashier_dashboard')
@login_required
@require_http_methods(["POST"])
def api_do_transfer(request):
    """JSON endpoint to perform a stock transfer."""
    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'JSON inválido.'}, status=400)
    try:
        producto_id = int(body.get('producto_id'))
        origen_id = int(body.get('origen_id'))
        destino_id = int(body.get('destino_id'))
        cantidad = int(body.get('cantidad'))
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Datos inválidos.'}, status=400)
    if origen_id == destino_id:
        return JsonResponse({'error': 'El origen y destino no pueden ser la misma sucursal.'}, status=400)
    if cantidad <= 0:
        return JsonResponse({'error': 'La cantidad debe ser mayor a cero.'}, status=400)
    producto = get_object_or_404(Product, id=producto_id)
    suc_origen = get_object_or_404(Sucursal, id=origen_id)
    suc_destino = get_object_or_404(Sucursal, id=destino_id)
    disp = producto.stock_en(suc_origen)
    allow_without_stock = producto.permitir_venta_sin_stock_en(suc_origen)
    if disp < cantidad and not allow_without_stock:
        return JsonResponse({'error': f'Stock insuficiente. Disponible: {disp}.'}, status=400)
    ss_origen, _ = StockSucursal.objects.get_or_create(producto=producto, sucursal=suc_origen, defaults={'cantidad': 0})
    ss_origen.cantidad = max(0, (ss_origen.cantidad or 0) - cantidad)
    ss_origen.save()
    ss_destino, _ = StockSucursal.objects.get_or_create(producto=producto, sucursal=suc_destino, defaults={'cantidad': 0})
    ss_destino.cantidad = (ss_destino.cantidad or 0) + cantidad
    ss_destino.save()
    TransferenciaStock.objects.create(
        producto=producto, origen=suc_origen, destino=suc_destino,
        cantidad=cantidad, usuario=request.user if request.user.is_authenticated else None
    )
    return JsonResponse({
        'ok': True,
        'message': f'{cantidad} unidades de "{producto.nombre}" transferidas de {suc_origen.nombre} a {suc_destino.nombre}.',
        'nuevo_stock_origen': ss_origen.cantidad,
        'nuevo_stock_destino': ss_destino.cantidad,
    })


@user_passes_test(_can_manage_products, login_url='cashier_dashboard')
@login_required
@require_http_methods(["GET"])
def api_transfer_history_json(request):
    """Return recent stock transfers as JSON."""
    try:
        limit = min(int(request.GET.get('limit', 10)), 50)
    except (ValueError, TypeError):
        limit = 10
    qs = TransferenciaStock.objects.select_related('producto', 'origen', 'destino', 'usuario').order_by('-fecha')[:limit]
    items = [{
        'id': t.id,
        'fecha': t.fecha.isoformat(),
        'producto': t.producto.nombre,
        'origen': t.origen.nombre,
        'destino': t.destino.nombre,
        'cantidad': t.cantidad,
        'usuario': t.usuario.username if t.usuario else None,
    } for t in qs]
    return JsonResponse({'items': items})


@user_passes_test(_is_staff, login_url='cashier_dashboard')
@login_required
def allow_without_stock_report(request):
    """Reporte rapido de overrides de venta sin stock por sucursal."""
    sucursales = Sucursal.objects.all().order_by('nombre')
    sucursal_id = request.GET.get('sucursal')
    q_text = (request.GET.get('q') or '').strip()

    qs = StockSucursal.objects.select_related('producto', 'sucursal')
    qs = qs.filter(permitir_venta_sin_stock__isnull=False)
    if sucursal_id:
        qs = qs.filter(sucursal_id=sucursal_id)
    if q_text:
        qs = qs.filter(
            Q(producto__nombre__icontains=q_text) |
            Q(producto__producto_id__icontains=q_text) |
            Q(producto__codigo_barras__icontains=q_text) |
            Q(sucursal__nombre__icontains=q_text)
        )

    qs = qs.order_by('sucursal__nombre', 'producto__nombre')

    return render(request, 'products/allow_without_stock_report.html', {
        'rows': qs,
        'sucursales': sucursales,
        'sucursal_sel': sucursal_id,
        'q_text': q_text,
    })

@user_passes_test(_can_manage_products, login_url='cashier_dashboard')
@login_required
def stock_availability(request):
    """Devuelve la cantidad disponible de un producto en una sucursal."""
    producto_id = request.GET.get('producto_id')
    sucursal_id = request.GET.get('sucursal_id')
    if not producto_id or not sucursal_id:
        return JsonResponse({'error': 'Producto y sucursal son obligatorios.'}, status=400)
    try:
        producto = Product.objects.get(id=int(producto_id))
        sucursal = Sucursal.objects.get(id=int(sucursal_id))
    except (Product.DoesNotExist, Sucursal.DoesNotExist, ValueError):
        return JsonResponse({'error': 'Producto o sucursal inválidos.'}, status=404)
    cantidad = producto.stock_en(sucursal)
    return JsonResponse({'producto_id': producto.id, 'sucursal_id': sucursal.id, 'cantidad': cantidad})


@user_passes_test(_can_manage_products, login_url='cashier_dashboard')
@login_required
def transfer_history(request):
    """Historial simple de transferencias con filtros básicos."""
    productos = Product.objects.all().order_by('nombre')
    sucursales = Sucursal.objects.all().order_by('nombre')
    qs = TransferenciaStock.objects.select_related('producto', 'origen', 'destino', 'usuario').all()
    prod = request.GET.get('producto')
    suc = request.GET.get('sucursal')
    # Paginación
    per_page_options = [10, 20, 30, 50]
    try:
        per_page = int(request.GET.get('per_page', 10))
    except ValueError:
        per_page = 10
    if per_page not in per_page_options:
        per_page = 10
    if prod:
        qs = qs.filter(producto_id=prod)
    if suc:
        qs = qs.filter(Q(origen_id=suc) | Q(destino_id=suc))
    # Ordenar descendente por fecha (más recientes primero)
    qs = qs.order_by('-fecha')
    # Construir paginador
    paginator = Paginator(qs, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'products/transfer_history.html', {
        'page_obj': page_obj,
        'per_page': per_page,
        'per_page_options': per_page_options,
        'productos': productos,
        'sucursales': sucursales,
        'producto_sel': prod,
        'sucursal_sel': suc,
    })

@user_passes_test(_can_manage_products, login_url='cashier_dashboard')
@login_required
def ajustar_stock(request):
    """Endpoint POST para ajustar stock por sucursal (incremento o decremento)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    try:
        producto_id = int(request.POST.get('producto_id'))
        sucursal_id = int(request.POST.get('sucursal_id'))
        delta = int(request.POST.get('delta'))
        motivo = (request.POST.get('motivo') or '').strip()
        producto = get_object_or_404(Product, id=producto_id)
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        ss, _ = StockSucursal.objects.get_or_create(producto=producto, sucursal=sucursal, defaults={'cantidad': 0})
        ss.cantidad = max(0, (ss.cantidad or 0) + delta)
        ss.save()
        AjusteStock.objects.create(
            producto=producto,
            sucursal=sucursal,
            cantidad_delta=delta,
            motivo=motivo or None,
            usuario=request.user if request.user.is_authenticated else None
        )
        return JsonResponse({'success': True, 'nueva_cantidad': ss.cantidad})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

def adjust_history(request):
    """Historial de ajustes de stock con filtros por producto y sucursal."""
    productos = Product.objects.all().order_by('nombre')
    sucursales = Sucursal.objects.all().order_by('nombre')
    qs = AjusteStock.objects.select_related('producto', 'sucursal', 'usuario').all()
    prod = request.GET.get('producto')
    suc = request.GET.get('sucursal')
    q_text = (request.GET.get('q') or '').strip()
    if prod:
        qs = qs.filter(producto_id=prod)
    if suc:
        qs = qs.filter(sucursal_id=suc)
    if q_text:
        qs = qs.filter(
            Q(producto__nombre__icontains=q_text) |
            Q(producto__producto_id__icontains=q_text) |
            Q(producto__codigo_alternativo__icontains=q_text)
        )
    # Ordenar por fecha descendente para mostrar primero los ajustes más recientes
    qs = qs.order_by('-fecha')
    return render(request, 'products/adjust_history.html', {
        'ajustes': qs[:200],
        'productos': productos,
        'sucursales': sucursales,
        'producto_sel': prod,
        'sucursal_sel': suc,
        'q_text': q_text,
    })

def sucursal_products(request, sucursal_id):
    """
    Listado de productos de una sucursal con paginación y tamaño de página configurable.
    """
    sucursal = get_object_or_404(Sucursal, id=sucursal_id)
    search_query = request.GET.get('search', '').strip()
    stock_filter = request.GET.get('stock', '').strip()  # '', 'low', 'out'
    try:
        per_page = int(request.GET.get('per_page', 10))
    except ValueError:
        per_page = 10
    if per_page not in [10, 20, 30, 50]:
        per_page = 10
    # Incluir productos asociados por FK y/o que tengan stock registrado en esta sucursal
    productos_fk = Product.objects.filter(sucursal_id=sucursal_id)
    productos_stock = Product.objects.filter(stocks_por_sucursal__sucursal_id=sucursal_id)
    productos_qs = (productos_fk | productos_stock).distinct().prefetch_related('stocks_por_sucursal')
    if search_query:
        productos_qs = productos_qs.filter(build_product_search_q(search_query))
    # Determinar umbral ahora para posible filtrado por stock
    threshold = None  # será resuelto abajo usando sucursal o settings
    # Filtrado por estado de stock si corresponde (antes de paginar)
    if stock_filter in ('low', 'out'):
        # Resolver threshold preliminar (si sucursal tiene 0, luego usaremos el global)
        threshold = sucursal.low_stock_threshold if getattr(sucursal, 'low_stock_threshold', 0) and sucursal.low_stock_threshold > 0 else getattr(settings, 'LOW_STOCK_THRESHOLD', 2)
        ids_keep = []
        for p in productos_qs:
            s_val = 0
            try:
                s_val = p.stock_en(sucursal)
            except Exception:
                s_val = p.stock or 0
            if stock_filter == 'out' and s_val <= 0:
                ids_keep.append(p.id)
            elif stock_filter == 'low' and s_val > 0 and s_val <= threshold:
                ids_keep.append(p.id)
        productos_qs = productos_qs.filter(id__in=ids_keep)
    productos_qs = productos_qs.order_by('nombre')
    paginator = Paginator(productos_qs, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # Adjuntar stock por sucursal a cada producto de la página
    for p in page_obj.object_list:
        try:
            p.stock_sucursal = p.stock_en(sucursal)
        except Exception:
            p.stock_sucursal = p.stock
        try:
            p.allow_without_stock = p.permitir_venta_sin_stock_en(sucursal)
        except Exception:
            p.allow_without_stock = p.permitir_venta_sin_stock
        ss = None
        try:
            ss = p.stocks_por_sucursal.filter(sucursal=sucursal).first()
        except Exception:
            ss = None
        if ss and ss.permitir_venta_sin_stock is not None:
            p.allow_override = ss.permitir_venta_sin_stock
        else:
            p.allow_override = None
    # Determinar umbral: usar el de la sucursal si > 0; en caso contrario, el global
    if threshold is None:
        threshold = sucursal.low_stock_threshold if getattr(sucursal, 'low_stock_threshold', 0) and sucursal.low_stock_threshold > 0 else getattr(settings, 'LOW_STOCK_THRESHOLD', 2)
    return render(request, 'sucursales/sucursal_products.html', {
        'sucursal': sucursal,
        'page_obj': page_obj,
        'per_page': per_page,
        'per_page_options': [10, 20, 30, 50],
        'search_query': search_query,
        'low_stock_threshold': threshold,
        'stock_filter': stock_filter
    })


@user_passes_test(_can_manage_products, login_url='cashier_dashboard')
@login_required
@require_http_methods(["GET"])
def api_products_list(request):
    """JSON API for React products management (search + paging + sorting)."""
    query = request.GET.get('search', '')
    sort_by = request.GET.get('sort_by', 'nombre')
    order = request.GET.get('order', 'asc')

    allowed_sort_fields = {
        'nombre': 'nombre',
        'descripcion': 'descripcion',
        'codigo1': 'producto_id',
        'codigo_barras': 'codigo_barras',
        'fecha_ingreso': 'fecha_ingreso_producto',
        'precio_compra': 'precio_compra',
        'precio_venta': 'precio_venta',
        'cantidad': 'cantidad',
        'stock': 'stock',
    }

    products = Product.objects.filter(build_product_search_q(query))
    if sort_by in allowed_sort_fields:
        field_to_sort = allowed_sort_fields[sort_by]
        if order == 'desc':
            field_to_sort = '-' + field_to_sort
        products = products.order_by(field_to_sort)
    else:
        products = products.order_by('nombre')

    per_page = request.GET.get('per_page', 10)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10
    per_page = max(1, min(200, per_page))

    page = request.GET.get('page', 1)
    paginator = Paginator(products, per_page)
    try:
        products_page = paginator.page(page)
    except PageNotAnInteger:
        products_page = paginator.page(1)
    except EmptyPage:
        products_page = paginator.page(paginator.num_pages)

    items = []
    for p in products_page.object_list:
        items.append({
            'id': p.id,
            'nombre': p.nombre or '',
            'producto_id': p.producto_id,
            'codigo_barras': p.codigo_barras or '',
            'precio_compra': str(p.precio_compra or '0'),
            'precio_venta': str(p.precio_venta or '0'),
            'stock': int(p.stock or 0),
            'cantidad': int(p.cantidad or 0),
            'permitir_venta_sin_stock': bool(p.permitir_venta_sin_stock),
            'activo': bool(getattr(p, 'activo', True)),
        })

    return JsonResponse({
        'items': items,
        'page': products_page.number,
        'num_pages': paginator.num_pages,
        'total': paginator.count,
        'per_page': per_page,
        'search': query,
        'sort_by': sort_by,
        'order': order,
    })