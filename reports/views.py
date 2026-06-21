import json as _json
import logging
import os
import math
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.utils.dateparse import parse_date
from django.db.models import Q, Sum, F, Count
from django.db.models.functions import Cast
from django.utils import timezone
import datetime
from decimal import Decimal, ROUND_HALF_UP
from django.http import JsonResponse, HttpResponse
from django.core.cache import cache
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_http_methods
from django.contrib.auth import get_user_model
from django.template.loader import render_to_string
import io
from django.views import View
from django.db.models import Avg
from django.db.models.functions import ExtractHour, ExtractWeekDay, ExtractDay, TruncDate, TruncMonth, TruncMinute
from django.conf import settings
from cashier.models import Venta, VentaDetalle, VentaPago, AperturaCierreCaja, Devolucion, DevolucionDetalle, NotaCredito
from cashier.views import format_clp
from MOVOS.money import to_clp_pesos
from sucursales.models import Sucursal
from products.models import Product
from django.db.models import DecimalField, ExpressionWrapper
from django.db import connection
from django.db.models import Case, When, Value, IntegerField, CharField
import calendar
from .models import Conversation, Message
from MOVOS.react_ui import react_ui_enabled, vite_dev_server_url_if_up

logger = logging.getLogger(__name__)
User = get_user_model()


def _clp_int(value) -> int:
    try:
        return int(to_clp_pesos(value))
    except Exception:
        return 0


def _clp_plain(value) -> str:
    """CLP formatted without currency symbol, no decimals."""
    try:
        return format_clp(to_clp_pesos(value))
    except Exception:
        return '0'


def _is_pct_key(key):
    if not key:
        return False
    k = key.lower()
    return ('pct' in k) or ('porcentaje' in k) or (k == 'margen') or k.endswith('_pct') or k.endswith('_rate')

def _local_assistant_response(prompt_text, context_summary):
    """Fallback simple assistant that crafts a short answer from context (used when no external model configured)."""
    # Simple heuristic: echo key KPIs and answer generically
    k = context_summary.get('kpis', {})
    inc = k.get('ingreso_total_clp')
    gan = k.get('ganancia_neta_clp')
    trans = k.get('num_transacciones')
    reply = (
        f"Resumen de datos: ingreso_total={inc}, ganancia_neta={gan}, transacciones={trans}.\n"
        + "Respuesta: "
        + (prompt_text[:800] or "Estoy listo para ayudar con tus KPIs.")
    )
    return reply


def _call_inference(prompt_text, context_summary):
    """Call a remote inference service if configured via settings env var REPORTS_INFERENCE_URL.
    If not configured, use local assistant fallback.
    """
    url = os.environ.get('REPORTS_INFERENCE_URL')
    if not url:
        return _local_assistant_response(prompt_text, context_summary)
    try:
        import requests

        payload = {
            'inputs': f"CONTEXT:\n{_json.dumps(context_summary)}\n\nUSER:\n{prompt_text}",
            'options': {'wait_for_model': True},
        }
        resp = requests.post(url, json=payload, timeout=15)
        if resp.ok:
            data = resp.json()
            # HF inference outputs can vary; try to extract 'generated_text' or first item
            if isinstance(data, dict) and 'generated_text' in data:
                return data['generated_text']
            if isinstance(data, list) and data and isinstance(data[0], dict):
                return data[0].get('generated_text') or str(data[0])
            return str(data)
        return f"(inference error {resp.status_code})"
    except Exception as e:
        return f"(inference exception) {e}"


def _is_admin(user):
    return user.is_authenticated and (
        user.is_staff
        or user.is_superuser
        or bool(getattr(user, 'can_view_analytics', False))
    )


def _parse_market_filters(request):
    """Parse market analysis filters from querystring.
    Accepts: start_date/end_date (preferred) and fecha_inicio/fecha_fin (legacy naming).
    Returns (start_dt, end_dt, sucursal_id).
    """
    start_str = (request.GET.get('start_date') or request.GET.get('fecha_inicio') or '').strip()
    end_str = (request.GET.get('end_date') or request.GET.get('fecha_fin') or '').strip()
    sucursal_id = (request.GET.get('sucursal_id') or request.GET.get('sucursal') or '').strip()

    start_date = parse_date(start_str) if start_str else None
    end_date = parse_date(end_str) if end_str else None

    start_dt = None
    end_dt = None

    if start_date:
        start_dt = timezone.make_aware(datetime.datetime.combine(start_date, datetime.time.min))
    if end_date:
        end_dt = timezone.make_aware(datetime.datetime.combine(end_date, datetime.time.max))

    return start_dt, end_dt, (int(sucursal_id) if sucursal_id.isdigit() else None)


def _ventas_filtered_qs(start_dt, end_dt, sucursal_id=None):
    qs = Venta.objects.all()
    if start_dt:
        qs = qs.filter(fecha__gte=start_dt)
    if end_dt:
        qs = qs.filter(fecha__lte=end_dt)
    if sucursal_id:
        qs = qs.filter(sucursal_id=sucursal_id)
    return qs


def _resolve_product_ref_to_pk(product_ref: str):
    """Resolve a product reference (pk or business codes) to Product.pk.

    Accepts:
    - internal pk (id)
    - codigo_barras
    - producto_id ("Código 1")
    - codigo_alternativo ("Código 2")
    """
    ref = (product_ref or '').strip()
    if not ref:
        return None

    # Prefer pk if it exists.
    if ref.isdigit():
        pk = int(ref)
        if Product.objects.filter(id=pk).exists():
            return pk

    prod = (
        Product.objects.filter(
            Q(codigo_barras=ref) | Q(producto_id=ref) | Q(codigo_alternativo=ref)
        )
        .only('id')
        .first()
    )
    return int(prod.id) if prod else None


def _require_applied_filters(start_dt, end_dt):
    return bool(start_dt and end_dt)


def _dt_range_for_dates(start_date, end_date):
    start_dt = timezone.make_aware(datetime.datetime.combine(start_date, datetime.time.min))
    end_dt = timezone.make_aware(datetime.datetime.combine(end_date, datetime.time.max))
    return start_dt, end_dt


def _safe_year_shift(dt, years_delta):
    if not dt:
        return None
    try:
        return dt.replace(year=dt.year + years_delta)
    except ValueError:
        # Feb 29 -> Feb 28 for non-leap years
        if dt.month == 2 and dt.day == 29:
            return dt.replace(year=dt.year + years_delta, day=28)
        raise


def _aggregate_sales_core(ventas_qs):
    agg = ventas_qs.aggregate(total_monto=Sum('total'), total_count=Count('id'))
    total_monto = agg.get('total_monto') or Decimal('0.00')
    total_count = int(agg.get('total_count') or 0)

    vd_qs = VentaDetalle.objects.filter(venta__in=ventas_qs).select_related('producto')

    total_items = vd_qs.aggregate(total=Sum('cantidad')).get('total') or 0
    total_items = int(total_items or 0)

    line_profit = ExpressionWrapper(
        (F('precio_unitario') - F('producto__precio_compra')) * F('cantidad'),
        output_field=DecimalField(max_digits=18, decimal_places=2)
    )
    gross_profit = vd_qs.aggregate(total=Sum(line_profit)).get('total') or Decimal('0.00')

    line_cost = ExpressionWrapper(
        F('producto__precio_compra') * F('cantidad'),
        output_field=DecimalField(max_digits=18, decimal_places=2)
    )
    total_cost = vd_qs.aggregate(total=Sum(line_cost)).get('total') or Decimal('0.00')

    # Estimated discounts: compare list price vs unit price (only when unit < list)
    discount_expr = ExpressionWrapper(
        (F('producto__precio_venta') - F('precio_unitario')) * F('cantidad'),
        output_field=DecimalField(max_digits=18, decimal_places=2)
    )
    total_discount = (
        vd_qs.filter(precio_unitario__lt=F('producto__precio_venta'))
        .aggregate(total=Sum(discount_expr))
        .get('total')
    ) or Decimal('0.00')

    avg_ticket = (total_monto / Decimal(total_count)) if total_count else Decimal('0.00')
    avg_items_per_ticket = (Decimal(total_items) / Decimal(total_count)) if total_count else Decimal('0.00')
    gross_margin_pct = (gross_profit / total_monto * Decimal('100.0')) if total_monto else Decimal('0.00')

    return {
        'total_monto': total_monto,
        'total_count': total_count,
        'avg_ticket': avg_ticket,
        'total_items': total_items,
        'avg_items_per_ticket': avg_items_per_ticket,
        'gross_profit': gross_profit,
        'total_cost': total_cost,
        'gross_margin_pct': gross_margin_pct,
        'total_discount_estimated': total_discount,
    }


def _avg_distinct_skus_per_ticket(ventas_qs):
    """Average distinct products per ticket (SKU breadth)."""
    vd = VentaDetalle.objects.filter(venta__in=ventas_qs)
    avg_val = (
        vd.values('venta_id')
        .annotate(distinct_skus=Count('producto_id', distinct=True))
        .aggregate(avg=Avg('distinct_skus'))
        .get('avg')
    )
    return float(avg_val or 0.0)


def _ventas_filtered_qs_with_employee(start_dt, end_dt, sucursal_id=None, empleado_id=None):
    qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    if empleado_id:
        qs = qs.filter(empleado_id=empleado_id)
    return qs


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_analysis(request):
    today = timezone.localdate()
    sucursales = [{'id': s.id, 'nombre': s.nombre} for s in Sucursal.objects.all().order_by('nombre')]
    ctx = {
        'react_context': {
            'default_start_date': (today - datetime.timedelta(days=30)).strftime('%Y-%m-%d'),
            'default_end_date': today.strftime('%Y-%m-%d'),
            'sucursales': sucursales,
        }
    }
    vite_url = vite_dev_server_url_if_up()
    if vite_url:
        ctx['vite_dev_server_url'] = vite_url
    return render(request, 'reports/react_market_analysis.html', ctx)


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_kpis_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)

    agg = ventas_qs.aggregate(
        total_monto=Sum('total'),
        total_count=Count('id'),
    )
    total_monto = agg.get('total_monto') or Decimal('0.00')
    total_count = agg.get('total_count') or 0

    ticket_promedio = (total_monto / Decimal(total_count)) if total_count else Decimal('0.00')

    # Margen total: sum((precio_unitario - costo) * cantidad)
    vd_qs = VentaDetalle.objects.filter(venta__in=ventas_qs).select_related('producto')
    line_profit = ExpressionWrapper(
        (F('precio_unitario') - F('producto__precio_compra')) * F('cantidad'),
        output_field=DecimalField(max_digits=18, decimal_places=2)
    )
    margen_total = vd_qs.aggregate(total=Sum(line_profit)).get('total') or Decimal('0.00')

    return JsonResponse({
        'filters': {
            'start_date': start_dt.date().isoformat() if start_dt else None,
            'end_date': end_dt.date().isoformat() if end_dt else None,
            'sucursal_id': sucursal_id,
        },
        'kpis': {
            'total_ventas_monto': _clp_int(total_monto),
            'total_ventas_count': int(total_count),
            'ticket_promedio': _clp_int(ticket_promedio),
            'margen_ganancia_total': _clp_int(margen_total),
        }
    })


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_overview_api(request):
    """High-level KPIs. Designed to be the first payload loaded after filters are applied."""
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    core = _aggregate_sales_core(ventas_qs)
    sku_breadth_avg = _avg_distinct_skus_per_ticket(ventas_qs)

    # Period windows: day/week/month ending at end_dt (local date boundaries)
    end_local_date = timezone.localdate(end_dt)
    day_start, day_end = _dt_range_for_dates(end_local_date, end_local_date)
    week_start_date = end_local_date - datetime.timedelta(days=6)
    week_start, week_end = _dt_range_for_dates(week_start_date, end_local_date)
    month_start_date = end_local_date.replace(day=1)
    month_start, month_end = _dt_range_for_dates(month_start_date, end_local_date)

    day_total = (_ventas_filtered_qs(day_start, day_end, sucursal_id=sucursal_id).aggregate(s=Sum('total')).get('s') or Decimal('0.00'))
    week_total = (_ventas_filtered_qs(week_start, week_end, sucursal_id=sucursal_id).aggregate(s=Sum('total')).get('s') or Decimal('0.00'))
    month_total = (_ventas_filtered_qs(month_start, month_end, sucursal_id=sucursal_id).aggregate(s=Sum('total')).get('s') or Decimal('0.00'))

    # Payment method breakdown
    payment_rows = (
        ventas_qs.values('forma_pago')
        .annotate(total_monto=Sum('total'), ventas_count=Count('id'))
        .order_by('-total_monto')
    )
    payment = [
        {
            'forma_pago': r.get('forma_pago') or '',
            'total_monto': _clp_int(r.get('total_monto') or Decimal('0.00')),
            'ventas_count': int(r.get('ventas_count') or 0),
        }
        for r in payment_rows
    ]

    # Returns / cancellations are not tracked in current schema
    returns_info = {
        'available': False,
        'rate_pct': None,
        'message': 'No hay un campo de devoluciones/anulaciones en el modelo actual.'
    }

    # Top product (single) by amount within the filtered range
    vd_qs = VentaDetalle.objects.filter(venta__in=ventas_qs).select_related('producto')
    line_amount = ExpressionWrapper(
        F('cantidad') * F('precio_unitario'),
        output_field=DecimalField(max_digits=18, decimal_places=2)
    )
    top_amount_row = (
        vd_qs.values('producto_id', 'producto__nombre')
        .annotate(unidades=Sum('cantidad'), total_monto=Sum(line_amount))
        .order_by('-total_monto', '-unidades')
        .first()
    )
    top_units_row = (
        vd_qs.values('producto_id', 'producto__nombre')
        .annotate(unidades=Sum('cantidad'), total_monto=Sum(line_amount))
        .order_by('-unidades', '-total_monto')
        .first()
    )

    def _row_to_top(row):
        if not row or not row.get('producto_id'):
            return None
        return {
            'producto_id': int(row.get('producto_id') or 0),
            'nombre': row.get('producto__nombre') or '',
            'unidades': int(row.get('unidades') or 0),
            'total_monto': _clp_int(row.get('total_monto') or Decimal('0.00')),
        }

    top_product = _row_to_top(top_amount_row)
    top_product_by_units = _row_to_top(top_units_row)

    return JsonResponse({
        'filters': {
            'start_date': start_dt.date().isoformat(),
            'end_date': end_dt.date().isoformat(),
            'sucursal_id': sucursal_id,
        },
        'periods': {
            'day': {'start': day_start.date().isoformat(), 'end': day_end.date().isoformat(), 'total_monto': _clp_int(day_total)},
            'week': {'start': week_start.date().isoformat(), 'end': week_end.date().isoformat(), 'total_monto': _clp_int(week_total)},
            'month': {'start': month_start.date().isoformat(), 'end': month_end.date().isoformat(), 'total_monto': _clp_int(month_total)},
        },
        'kpis': {
            'venta_total_neta': _clp_int(core['total_monto']),
            'ventas_count': int(core['total_count']),
            'utilidad_bruta_real': _clp_int(core['gross_profit']),
            'costo_total_estimado': _clp_int(core['total_cost']),
            'margen_ganancia_promedio_pct': float(core['gross_margin_pct']),
            'ticket_promedio': _clp_int(core['avg_ticket']),
            'articulos_por_ticket': float(core['avg_items_per_ticket']),
            'sku_breadth_avg': float(sku_breadth_avg),
            'total_descuentos_estimado': _clp_int(core['total_discount_estimated']),
        },
        'top_product': top_product,
        'top_product_by_units': top_product_by_units,
        'payment_methods': payment,
        'returns': returns_info,
    })


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_heatmap_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    rows = (
        ventas_qs.annotate(weekday=ExtractWeekDay('fecha'), hour=ExtractHour('fecha'))
        .values('weekday', 'hour')
        .annotate(total_monto=Sum('total'), ventas_count=Count('id'))
        .order_by('weekday', 'hour')
    )

    matrix_amount = [[0 for _ in range(24)] for _ in range(7)]
    matrix_count = [[0 for _ in range(24)] for _ in range(7)]

    # ExtractWeekDay: Sunday=1 ... Saturday=7 (SQLite/Postgres)
    def idx_from_weekday(wd):
        if not wd:
            return None
        wd = int(wd)
        # convert to Monday=0 ... Sunday=6
        # Sunday(1)->6, Monday(2)->0, ..., Saturday(7)->5
        return 6 if wd == 1 else (wd - 2)

    for r in rows:
        wd = r.get('weekday')
        h = r.get('hour')
        i = idx_from_weekday(wd)
        if i is None or h is None:
            continue
        h = int(h)
        if 0 <= i <= 6 and 0 <= h <= 23:
            matrix_amount[i][h] = _clp_int(r.get('total_monto') or Decimal('0.00'))
            matrix_count[i][h] = int(r.get('ventas_count') or 0)

    # Peak hour by tickets (not amount)
    peak = (
        ventas_qs.annotate(hour=ExtractHour('fecha'))
        .values('hour')
        .annotate(ventas_count=Count('id'), total_monto=Sum('total'))
        .order_by('-ventas_count', '-total_monto')
        .first()
    )
    peak_hour = int(peak.get('hour')) if peak and peak.get('hour') is not None else None
    peak_tickets = int(peak.get('ventas_count') or 0) if peak else 0
    peak_amount = _clp_int(peak.get('total_monto') or Decimal('0.00')) if peak else 0

    peak_amount_row = (
        ventas_qs.annotate(hour=ExtractHour('fecha'))
        .values('hour')
        .annotate(total_monto=Sum('total'), ventas_count=Count('id'))
        .order_by('-total_monto', '-ventas_count')
        .first()
    )
    peak_hour_amount = int(peak_amount_row.get('hour')) if peak_amount_row and peak_amount_row.get('hour') is not None else None
    peak_amount_amount = _clp_int(peak_amount_row.get('total_monto') or Decimal('0.00')) if peak_amount_row else 0
    peak_tickets_amount = int(peak_amount_row.get('ventas_count') or 0) if peak_amount_row else 0

    return JsonResponse({
        'matrix_amount': matrix_amount,
        'matrix_count': matrix_count,
        'peak_hour': peak_hour,
        'peak_tickets': peak_tickets,
        'peak_amount': peak_amount,
        'peak_hour_amount': peak_hour_amount,
        'peak_amount_amount': peak_amount_amount,
        'peak_tickets_amount': peak_tickets_amount,
        'weekday_labels': ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom'],
    })


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_branch_comparison_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=None)  # compare across branches
    if sucursal_id:
        ventas_qs = ventas_qs.filter(sucursal_id=sucursal_id)

    sales_rows = (
        ventas_qs.values('sucursal_id', 'sucursal__nombre')
        .annotate(total_monto=Sum('total'), ventas_count=Count('id'))
        .order_by('-total_monto')
    )

    vd_qs = VentaDetalle.objects.filter(venta__in=ventas_qs)
    items_rows = (
        vd_qs.values('venta__sucursal_id')
        .annotate(total_items=Sum('cantidad'))
    )
    items_map = {int(r.get('venta__sucursal_id') or 0): int(r.get('total_items') or 0) for r in items_rows}

    out = []
    for r in sales_rows:
        sid = int(r.get('sucursal_id') or 0)
        total_monto = r.get('total_monto') or Decimal('0.00')
        count = int(r.get('ventas_count') or 0)
        items = int(items_map.get(sid, 0) or 0)
        avg_ticket = (total_monto / Decimal(count)) if count else Decimal('0.00')
        avg_items = (Decimal(items) / Decimal(count)) if count else Decimal('0.00')
        out.append({
            'sucursal_id': sid,
            'sucursal_nombre': r.get('sucursal__nombre') or '',
            'total_monto': _clp_int(total_monto),
            'ventas_count': count,
            'ticket_promedio': _clp_int(avg_ticket),
            'articulos_por_ticket': float(avg_items),
        })

    return JsonResponse({'branches': out})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_weekend_vs_weekday_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    rows = (
        ventas_qs.annotate(weekday=ExtractWeekDay('fecha'))
        .values('weekday')
        .annotate(total_monto=Sum('total'), ventas_count=Count('id'))
    )

    weekend_total = Decimal('0.00')
    weekend_count = 0
    weekday_total = Decimal('0.00')
    weekday_count = 0

    for r in rows:
        wd = int(r.get('weekday') or 0)
        tot = r.get('total_monto') or Decimal('0.00')
        cnt = int(r.get('ventas_count') or 0)
        # Sunday=1, Saturday=7
        if wd in (1, 7):
            weekend_total += tot
            weekend_count += cnt
        else:
            weekday_total += tot
            weekday_count += cnt

    return JsonResponse({
        'weekend': {'total_monto': _clp_int(weekend_total), 'ventas_count': int(weekend_count)},
        'weekday': {'total_monto': _clp_int(weekday_total), 'ventas_count': int(weekday_count)},
    })


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_monthly_trend_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    rows = (
        ventas_qs.annotate(month=TruncMonth('fecha'))
        .values('month')
        .annotate(total_monto=Sum('total'), ventas_count=Count('id'))
        .order_by('month')
    )
    out = []
    prev = None
    for r in rows:
        month_dt = r.get('month')
        total = r.get('total_monto') or Decimal('0.00')
        growth_pct = None
        if prev is not None and prev != Decimal('0.00'):
            growth_pct = float(((total - prev) / prev) * Decimal('100.0'))
        out.append({
            'month': month_dt.date().isoformat() if month_dt else None,
            'total_monto': _clp_int(total),
            'ventas_count': int(r.get('ventas_count') or 0),
            'growth_pct': growth_pct,
        })
        prev = total
    return JsonResponse({'monthly': out})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_top_margin_products_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    vd_qs = VentaDetalle.objects.filter(venta__in=ventas_qs)

    monto_expr = ExpressionWrapper(
        F('cantidad') * F('precio_unitario'),
        output_field=DecimalField(max_digits=18, decimal_places=2)
    )
    profit_expr = ExpressionWrapper(
        (F('precio_unitario') - F('producto__precio_compra')) * F('cantidad'),
        output_field=DecimalField(max_digits=18, decimal_places=2)
    )

    base = (
        vd_qs.values('producto_id', 'producto__nombre')
        .annotate(
            unidades=Sum('cantidad'),
            ingreso=Sum(monto_expr),
            utilidad=Sum(profit_expr),
        )
    )

    rows = list(base)
    enriched = []
    for r in rows:
        ingreso = r.get('ingreso') or Decimal('0.00')
        utilidad = r.get('utilidad') or Decimal('0.00')
        margen_pct = float((utilidad / ingreso) * Decimal('100.0')) if ingreso else 0.0
        enriched.append({
            'producto_id': r.get('producto_id'),
            'producto_nombre': r.get('producto__nombre') or '',
            'unidades': int(r.get('unidades') or 0),
            'ingreso': _clp_int(ingreso),
            'utilidad': _clp_int(utilidad),
            'margen_pct': margen_pct,
        })

    top_by_margin = sorted(enriched, key=lambda x: (x.get('margen_pct') or 0.0, x.get('utilidad') or 0.0), reverse=True)[:10]
    top_by_profit = sorted(enriched, key=lambda x: (x.get('utilidad') or 0.0, x.get('ingreso') or 0.0), reverse=True)[:10]

    return JsonResponse({'top_by_margin_pct': top_by_margin, 'top_by_profit': top_by_profit})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_zombie_products_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    days_raw = (request.GET.get('days') or '').strip()
    days = int(days_raw) if days_raw.isdigit() else 30
    days = max(1, min(days, 365))

    end_local_date = timezone.localdate(end_dt)
    start_date = end_local_date - datetime.timedelta(days=days)
    window_start, window_end = _dt_range_for_dates(start_date, end_local_date)

    sold_qs = VentaDetalle.objects.filter(venta__fecha__gte=window_start, venta__fecha__lte=window_end)
    if sucursal_id:
        sold_qs = sold_qs.filter(venta__sucursal_id=sucursal_id)

    sold_ids = sold_qs.values_list('producto_id', flat=True).distinct()

    products_qs = Product.objects.all().order_by('nombre')
    if sucursal_id:
        products_qs = products_qs.filter(sucursal_id=sucursal_id)

    zombies = products_qs.exclude(id__in=sold_ids).values('id', 'nombre', 'stock', 'sucursal_id')[:50]
    out = []
    for p in zombies:
        out.append({
            'producto_id': int(p.get('id')),
            'producto_nombre': p.get('nombre') or '',
            'stock': int(p.get('stock') or 0),
            'sucursal_id': p.get('sucursal_id'),
        })
    return JsonResponse({'days': days, 'zombies': out})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_pareto_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    limit_raw = (request.GET.get('limit') or '').strip()
    limit = int(limit_raw) if limit_raw.isdigit() else 30
    limit = max(5, min(limit, 100))

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    vd_qs = VentaDetalle.objects.filter(venta__in=ventas_qs)
    monto_expr = ExpressionWrapper(F('cantidad') * F('precio_unitario'), output_field=DecimalField(max_digits=18, decimal_places=2))

    rows = list(
        vd_qs.values('producto_id', 'producto__nombre')
        .annotate(monto=Sum(monto_expr))
        .order_by('-monto')[:limit]
    )
    total = (vd_qs.aggregate(t=Sum(monto_expr)).get('t') or Decimal('0.00'))
    cum = Decimal('0.00')
    out = []
    for i, r in enumerate(rows, start=1):
        monto = r.get('monto') or Decimal('0.00')
        cum += monto
        out.append({
            'rank': i,
            'producto_id': r.get('producto_id'),
            'producto_nombre': r.get('producto__nombre') or '',
            'monto': _clp_int(monto),
            'cum_share_pct': float((cum / total) * Decimal('100.0')) if total else 0.0,
        })
    return JsonResponse({'total_monto': _clp_int(total), 'series': out})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_price_distribution_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    vd_qs = VentaDetalle.objects.filter(venta__in=ventas_qs)

    monto_expr = ExpressionWrapper(F('cantidad') * F('precio_unitario'), output_field=DecimalField(max_digits=18, decimal_places=2))
    # We use an int code for ordering + a human label
    bucket = Case(
        When(precio_unitario__lt=1000, then=Value(0)),
        When(precio_unitario__lt=5000, then=Value(1)),
        When(precio_unitario__lt=10000, then=Value(2)),
        When(precio_unitario__lt=20000, then=Value(3)),
        default=Value(4),
        output_field=IntegerField(),
    )
    label = Case(
        When(precio_unitario__lt=1000, then=Value('0–999')),
        When(precio_unitario__lt=5000, then=Value('1.000–4.999')),
        When(precio_unitario__lt=10000, then=Value('5.000–9.999')),
        When(precio_unitario__lt=20000, then=Value('10.000–19.999')),
        default=Value('20.000+'),
        output_field=CharField(),
    )
    rows = (
        vd_qs.annotate(bucket=bucket, label=label)
        .values('bucket', 'label')
        .annotate(items=Sum('cantidad'), revenue=Sum(monto_expr))
        .order_by('bucket')
    )
    out = []
    for r in rows:
        out.append({
            'label': r.get('label') or '',
            'items': int(r.get('items') or 0),
            'revenue': _clp_int(r.get('revenue') or Decimal('0.00')),
        })
    return JsonResponse({'buckets': out})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_offers_vs_full_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    vd_qs = VentaDetalle.objects.filter(venta__in=ventas_qs).select_related('producto')
    monto_expr = ExpressionWrapper(F('cantidad') * F('precio_unitario'), output_field=DecimalField(max_digits=18, decimal_places=2))

    offer_qs = vd_qs.filter(precio_unitario__lt=F('producto__precio_venta'))
    full_qs = vd_qs.exclude(precio_unitario__lt=F('producto__precio_venta'))

    def agg(qs):
        a = qs.aggregate(items=Sum('cantidad'), revenue=Sum(monto_expr))
        return {
            'items': int(a.get('items') or 0),
            'revenue': _clp_int(a.get('revenue') or Decimal('0.00')),
        }

    return JsonResponse({'offer': agg(offer_qs), 'full': agg(full_qs)})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_impulse_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    impulse_max_raw = (request.GET.get('impulse_max') or '').strip()
    ticket_min_raw = (request.GET.get('ticket_min') or '').strip()
    impulse_max = Decimal(impulse_max_raw) if impulse_max_raw.replace('.', '', 1).isdigit() else Decimal('2000')
    ticket_min = Decimal(ticket_min_raw) if ticket_min_raw.replace('.', '', 1).isdigit() else Decimal('10000')

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id).filter(total__gte=ticket_min)
    vd_qs = VentaDetalle.objects.filter(venta__in=ventas_qs, precio_unitario__lte=impulse_max)

    monto_expr = ExpressionWrapper(F('cantidad') * F('precio_unitario'), output_field=DecimalField(max_digits=18, decimal_places=2))
    rows = (
        vd_qs.values('producto_id', 'producto__nombre')
        .annotate(items=Sum('cantidad'), revenue=Sum(monto_expr))
        .order_by('-items', '-revenue')[:10]
    )
    out = []
    for r in rows:
        out.append({
            'producto_id': r.get('producto_id'),
            'producto_nombre': r.get('producto__nombre') or '',
            'items': int(r.get('items') or 0),
            'revenue': _clp_int(r.get('revenue') or Decimal('0.00')),
        })

    return JsonResponse({'impulse_max': _clp_int(impulse_max), 'ticket_min': _clp_int(ticket_min), 'top_impulse': out})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_projection_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    end_local_date = timezone.localdate(end_dt)
    month_start_date = end_local_date.replace(day=1)
    mtd_start, mtd_end = _dt_range_for_dates(month_start_date, end_local_date)
    ventas_mtd = _ventas_filtered_qs(mtd_start, mtd_end, sucursal_id=sucursal_id)

    total_mtd = ventas_mtd.aggregate(s=Sum('total')).get('s') or Decimal('0.00')
    day_of_month = int(end_local_date.day)
    days_in_month = int(calendar.monthrange(end_local_date.year, end_local_date.month)[1])
    projection = (total_mtd / Decimal(day_of_month) * Decimal(days_in_month)) if day_of_month else Decimal('0.00')

    return JsonResponse({
        'month': end_local_date.strftime('%Y-%m'),
        'mtd_total': _clp_int(total_mtd),
        'day_of_month': day_of_month,
        'days_in_month': days_in_month,
        'projection_close_month': _clp_int(projection),
    })


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_yoy_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    prev_start = _safe_year_shift(start_dt, -1)
    prev_end = _safe_year_shift(end_dt, -1)

    ventas_now = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    ventas_prev = _ventas_filtered_qs(prev_start, prev_end, sucursal_id=sucursal_id)

    now_total = ventas_now.aggregate(s=Sum('total')).get('s') or Decimal('0.00')
    prev_total = ventas_prev.aggregate(s=Sum('total')).get('s') or Decimal('0.00')
    yoy_pct = float(((now_total - prev_total) / prev_total) * Decimal('100.0')) if prev_total else None

    return JsonResponse({
        'current': {'start': start_dt.date().isoformat(), 'end': end_dt.date().isoformat(), 'total_monto': _clp_int(now_total)},
        'prev_year': {'start': prev_start.date().isoformat() if prev_start else None, 'end': prev_end.date().isoformat() if prev_end else None, 'total_monto': _clp_int(prev_total)},
        'yoy_pct': yoy_pct,
    })


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_runway_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    producto_ref = (request.GET.get('producto_id') or '').strip()
    producto_id = _resolve_product_ref_to_pk(producto_ref)
    if not producto_id:
        return JsonResponse({'error': 'Producto inválido. Usa ID interno, código o código de barras.'}, status=400)

    days_raw = (request.GET.get('days') or '').strip()
    days = int(days_raw) if days_raw.isdigit() else 30
    days = max(1, min(days, 365))

    end_local_date = timezone.localdate(end_dt)
    window_start_date = end_local_date - datetime.timedelta(days=days)
    window_start, window_end = _dt_range_for_dates(window_start_date, end_local_date)

    ventas_window = _ventas_filtered_qs(window_start, window_end, sucursal_id=sucursal_id)
    sold = (
        VentaDetalle.objects.filter(venta__in=ventas_window, producto_id=producto_id)
        .aggregate(units=Sum('cantidad'))
        .get('units')
    ) or 0
    sold_units = Decimal(int(sold or 0))
    avg_daily_units = (sold_units / Decimal(days)) if days else Decimal('0.00')

    try:
        product = Product.objects.get(id=producto_id)
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Producto no existe.'}, status=404)

    stock = Decimal(int(product.stock or 0))
    if sucursal_id:
        suc = Sucursal.objects.filter(id=sucursal_id).first()
        if suc:
            try:
                stock = Decimal(int(product.stock_en(suc)))
            except Exception:
                stock = Decimal(int(product.stock or 0))

    runway_days = float((stock / avg_daily_units)) if avg_daily_units and avg_daily_units > 0 else None

    return JsonResponse({
        'producto_id': producto_id,
        'producto_nombre': product.nombre or '',
        'days_window': days,
        'stock': float(stock),
        'sold_units_window': float(sold_units),
        'avg_daily_units': float(avg_daily_units),
        'runway_days': runway_days,
    })


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_top_products_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    vd_qs = VentaDetalle.objects.filter(venta__in=ventas_qs)

    monto_expr = ExpressionWrapper(
        F('cantidad') * F('precio_unitario'),
        output_field=DecimalField(max_digits=18, decimal_places=2)
    )
    costo_expr = ExpressionWrapper(
        F('cantidad') * F('producto__precio_compra'),
        output_field=DecimalField(max_digits=18, decimal_places=2)
    )
    profit_expr = ExpressionWrapper(
        (F('precio_unitario') - F('producto__precio_compra')) * F('cantidad'),
        output_field=DecimalField(max_digits=18, decimal_places=2)
    )

    # NOTE: Do NOT annotate with name "cantidad" because it collides with the model
    # field name used inside ExpressionWrapper(F('cantidad') * ...), causing Django
    # to treat it as an aggregate (FieldError).
    base = (
        vd_qs.values('producto_id', 'producto__nombre')
        .annotate(
            unidades=Sum('cantidad'),
            monto_total=Sum(monto_expr),
            costo_total=Sum(costo_expr),
            ganancia_total=Sum(profit_expr),
        )
    )

    limit_raw = (request.GET.get('limit') or '').strip()
    limit = int(limit_raw) if limit_raw.isdigit() else 10
    limit = max(5, min(limit, 50))

    top_by_qty = list(base.order_by('-unidades', '-monto_total')[:limit])
    top_by_amount = list(base.order_by('-monto_total', '-unidades')[:limit])

    def norm(rows):
        out = []
        for r in rows:
            monto = r.get('monto_total') or Decimal('0.00')
            gan = r.get('ganancia_total') or Decimal('0.00')
            margen_pct = float((gan / monto) * Decimal('100.0')) if monto else 0.0
            out.append({
                'producto_id': r.get('producto_id'),
                'producto_nombre': r.get('producto__nombre') or '',
                'cantidad': int(r.get('unidades') or 0),
                'monto_total': _clp_int(monto),
                'costo_total': _clp_int(r.get('costo_total') or Decimal('0.00')),
                'ganancia_total': _clp_int(gan),
                'margen_pct': margen_pct,
            })
        return out

    return JsonResponse({
        'top_by_qty': norm(top_by_qty),
        'top_by_amount': norm(top_by_amount),
    })


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_sales_by_hour_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    rows = (
        ventas_qs.annotate(hour=ExtractHour('fecha'))
        .values('hour')
        .annotate(total_monto=Sum('total'), ventas_count=Count('id'))
        .order_by('hour')
    )
    out = []
    for r in rows:
        out.append({
            'hour': int(r.get('hour')) if r.get('hour') is not None else None,
            'total_monto': _clp_int(r.get('total_monto') or Decimal('0.00')),
            'ventas_count': int(r.get('ventas_count') or 0),
        })
    return JsonResponse({'by_hour': out})


def _market_basket_raw_sql(producto_id, start_dt, end_dt, sucursal_id=None, limit=5):
    """Return top co-occurring products with the given producto_id.

    Uses raw SQL for efficiency:
    - Finds ventas that contain producto_id
    - Counts distinct ventas where each other product appears alongside
    """
    vd_table = VentaDetalle._meta.db_table
    v_table = Venta._meta.db_table
    p_table = Product._meta.db_table

    where_parts = [
        'vd1.producto_id = %s',
        'vd2.producto_id != %s',
        'v.id = vd1.venta_id',
        'v.id = vd2.venta_id',
        'v.fecha >= %s',
        'v.fecha <= %s',
    ]
    params = [producto_id, producto_id, start_dt, end_dt]
    if sucursal_id:
        where_parts.append('v.sucursal_id = %s')
        params.append(sucursal_id)

    sql = f'''
        SELECT
            vd2.producto_id AS producto_id,
            p.nombre AS producto_nombre,
            COUNT(DISTINCT vd2.venta_id) AS cooc_sales,
            SUM(vd2.cantidad) AS cooc_qty
        FROM {vd_table} vd1
        JOIN {vd_table} vd2 ON vd1.venta_id = vd2.venta_id
        JOIN {v_table} v ON v.id = vd1.venta_id
        JOIN {p_table} p ON p.id = vd2.producto_id
        WHERE {' AND '.join(where_parts)}
        GROUP BY vd2.producto_id, p.nombre
        ORDER BY cooc_sales DESC, cooc_qty DESC
        LIMIT %s
    '''
    with connection.cursor() as cursor:
        cursor.execute(sql, params + [int(limit)])
        cols = [c[0] for c in cursor.description]
        rows = [dict(zip(cols, row)) for row in cursor.fetchall()]
    return rows


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_basket_api(request):
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    producto_ref = (request.GET.get('producto_id') or '').strip()
    producto_id = _resolve_product_ref_to_pk(producto_ref)
    if not producto_id:
        return JsonResponse({'error': 'Producto inválido. Usa ID interno, código o código de barras.'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    total_sales = ventas_qs.count()
    if total_sales <= 0:
        return JsonResponse({
            'total_sales': 0,
            'sales_with_product': 0,
            'recommendations': [],
        })

    # Sales containing A
    sales_with_a = (
        VentaDetalle.objects.filter(venta__in=ventas_qs, producto_id=producto_id)
        .values('venta_id').distinct().count()
    )

    raw_rows = _market_basket_raw_sql(producto_id, start_dt, end_dt, sucursal_id=sucursal_id, limit=5)

    # Enrich with confidence + lift
    # confidence(A->B) = cooc_sales / sales_with_a
    # lift(A->B) = confidence / (sales_with_b / total_sales)
    recommendations = []
    for r in raw_rows:
        b_id = int(r.get('producto_id'))
        cooc_sales = int(r.get('cooc_sales') or 0)
        sales_with_b = (
            VentaDetalle.objects.filter(venta__in=ventas_qs, producto_id=b_id)
            .values('venta_id').distinct().count()
        )
        confidence = (cooc_sales / float(sales_with_a)) if sales_with_a else 0.0
        base_rate_b = (sales_with_b / float(total_sales)) if total_sales else 0.0
        lift = (confidence / base_rate_b) if base_rate_b else 0.0

        recommendations.append({
            'producto_id': b_id,
            'producto_nombre': r.get('producto_nombre') or '',
            'cooc_sales': cooc_sales,
            'cooc_qty': int(r.get('cooc_qty') or 0),
            'sales_with_b': int(sales_with_b),
            'confidence': confidence,
            'lift': lift,
        })

    return JsonResponse({
        'total_sales': int(total_sales),
        'sales_with_product': int(sales_with_a),
        'recommendations': recommendations,
    })


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_attach_rate_api(request):
    """Attach rate = P(B|A) for top products B, given a primary product A."""
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    primary_ref = (request.GET.get('primary_id') or '').strip()
    primary_id = _resolve_product_ref_to_pk(primary_ref)
    if not primary_id:
        return JsonResponse({'error': 'Producto A inválido. Usa ID interno, código o código de barras.'}, status=400)

    limit_raw = (request.GET.get('limit') or '10').strip()
    limit = int(limit_raw) if limit_raw.isdigit() else 10
    limit = max(1, min(limit, 30))

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    tickets_with_a = (
        VentaDetalle.objects.filter(venta__in=ventas_qs, producto_id=primary_id)
        .values('venta_id').distinct().count()
    )
    if tickets_with_a <= 0:
        return JsonResponse({'primary_id': primary_id, 'tickets_with_primary': 0, 'attached': []})

    raw_rows = _market_basket_raw_sql(primary_id, start_dt, end_dt, sucursal_id=sucursal_id, limit=limit)
    attached = []
    for r in raw_rows:
        cooc_sales = int(r.get('cooc_sales') or 0)
        attach_rate = (cooc_sales / float(tickets_with_a)) if tickets_with_a else 0.0
        attached.append({
            'producto_id': int(r.get('producto_id')),
            'producto_nombre': r.get('producto_nombre') or '',
            'tickets_with_primary': int(tickets_with_a),
            'tickets_with_both': cooc_sales,
            'attach_rate': float(attach_rate),
        })

    return JsonResponse({'primary_id': primary_id, 'tickets_with_primary': int(tickets_with_a), 'attached': attached})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_antisocial_products_api(request):
    """Products frequently bought alone (single-product tickets)."""
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    limit_raw = (request.GET.get('limit') or '20').strip()
    limit = int(limit_raw) if limit_raw.isdigit() else 20
    limit = max(1, min(limit, 50))

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    vd_qs = VentaDetalle.objects.filter(venta__in=ventas_qs)

    solo_venta_ids = (
        vd_qs.values('venta_id')
        .annotate(distinct_products=Count('producto_id', distinct=True))
        .filter(distinct_products=1)
        .values('venta_id')
    )

    solo_rows = list(
        vd_qs.filter(venta_id__in=solo_venta_ids)
        .values('producto_id', 'producto__nombre')
        .annotate(solo_tickets=Count('venta_id', distinct=True))
        .order_by('-solo_tickets')[: max(limit * 3, limit)]
    )
    if not solo_rows:
        return JsonResponse({'antisocial': []})

    prod_ids = [int(r['producto_id']) for r in solo_rows]
    totals = {
        int(r['producto_id']): int(r['tickets'] or 0)
        for r in vd_qs.filter(producto_id__in=prod_ids)
        .values('producto_id')
        .annotate(tickets=Count('venta_id', distinct=True))
    }

    out = []
    for r in solo_rows:
        pid = int(r['producto_id'])
        solo_tickets = int(r['solo_tickets'] or 0)
        tickets = int(totals.get(pid) or 0)
        rate = (solo_tickets / float(tickets)) if tickets else 0.0
        out.append({
            'producto_id': pid,
            'producto_nombre': r.get('producto__nombre') or '',
            'solo_tickets': solo_tickets,
            'tickets_with_product': tickets,
            'solo_rate': float(rate),
        })

    out.sort(key=lambda x: (x['solo_rate'], x['solo_tickets']), reverse=True)
    return JsonResponse({'antisocial': out[:limit]})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_paycycle_api(request):
    """Compare sales on common paydays (1-5 and 15-20) vs the rest."""
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    annotated = ventas_qs.annotate(day=ExtractDay('fecha'))
    payday = annotated.filter(Q(day__gte=1, day__lte=5) | Q(day__gte=15, day__lte=20))
    rest = annotated.exclude(Q(day__gte=1, day__lte=5) | Q(day__gte=15, day__lte=20))

    p = payday.aggregate(total_monto=Sum('total'), tickets=Count('id'))
    r = rest.aggregate(total_monto=Sum('total'), tickets=Count('id'))

    return JsonResponse({
        'paydays': {
            'total_monto': _clp_int(p.get('total_monto') or 0),
            'tickets': int(p.get('tickets') or 0),
        },
        'rest': {
            'total_monto': _clp_int(r.get('total_monto') or 0),
            'tickets': int(r.get('tickets') or 0),
        },
    })


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_sales_by_terminal_api(request):
    """Sales grouped by terminal/caja (AperturaCierreCaja)."""
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    limit_raw = (request.GET.get('limit') or '10').strip()
    limit = int(limit_raw) if limit_raw.isdigit() else 10
    limit = max(1, min(limit, 30))

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    rows = list(
        ventas_qs.values('caja_id', 'caja__sucursal__nombre', 'caja__vendedor__username')
        .annotate(total_monto=Sum('total'), tickets=Count('id'))
        .order_by('-total_monto')[:limit]
    )
    out = []
    for r in rows:
        out.append({
            'caja_id': r.get('caja_id'),
            'caja_label': f"Caja {r.get('caja_id')}" if r.get('caja_id') else 'Sin caja',
            'sucursal_nombre': r.get('caja__sucursal__nombre') or '',
            'vendedor': r.get('caja__vendedor__username') or '',
            'total_monto': _clp_int(r.get('total_monto') or 0),
            'tickets': int(r.get('tickets') or 0),
        })
    return JsonResponse({'terminals': out})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_peak_density_api(request):
    """Max tickets per minute in the busiest hour."""
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    by_hour = list(
        ventas_qs.annotate(hour=ExtractHour('fecha'))
        .values('hour')
        .annotate(tickets=Count('id'))
        .order_by('-tickets')
    )
    if not by_hour:
        return JsonResponse({'peak_hour': None, 'max_tickets_per_minute': None})

    peak_hour = by_hour[0].get('hour')
    if peak_hour is None:
        return JsonResponse({'peak_hour': None, 'max_tickets_per_minute': None})

    by_minute = list(
        ventas_qs.filter(fecha__hour=int(peak_hour))
        .annotate(minute=TruncMinute('fecha'))
        .values('minute')
        .annotate(tickets=Count('id'))
        .order_by('-tickets')
    )
    max_tpm = int(by_minute[0].get('tickets') or 0) if by_minute else 0

    return JsonResponse({'peak_hour': int(peak_hour), 'max_tickets_per_minute': int(max_tpm)})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_demand_volatility_api(request):
    """Demand volatility by product using coefficient of variation over daily units."""
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    vd = VentaDetalle.objects.filter(venta__in=ventas_qs)

    daily = list(
        vd.annotate(day=TruncDate('venta__fecha'))
        .values('producto_id', 'producto__nombre', 'day')
        .annotate(units=Sum('cantidad'))
        .order_by('producto_id', 'day')
    )
    by_product = {}
    for r in daily:
        pid = int(r['producto_id'])
        by_product.setdefault(pid, {'producto_nombre': r.get('producto__nombre') or '', 'series': []})
        by_product[pid]['series'].append(float(r.get('units') or 0))

    stats = []
    for pid, data in by_product.items():
        series = data['series']
        if len(series) < 5:
            continue
        total_units = sum(series)
        if total_units < 10:
            continue
        mean = total_units / float(len(series))
        if mean <= 0:
            continue
        var = sum((x - mean) ** 2 for x in series) / float(len(series))
        std = math.sqrt(var)
        cv = std / mean if mean else 0.0
        stats.append({
            'producto_id': int(pid),
            'producto_nombre': data['producto_nombre'],
            'cv': float(cv),
            'units': float(total_units),
        })

    stats.sort(key=lambda x: x['cv'], reverse=True)
    high = stats[:10]
    low = sorted(stats, key=lambda x: x['cv'])[:10]

    return JsonResponse({'high': high, 'low': low})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_price_elasticity_api(request):
    """Historical price elasticity approximation for a given product, using month buckets."""
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    producto_ref = (request.GET.get('producto_id') or '').strip()
    producto_id = _resolve_product_ref_to_pk(producto_ref)
    if not producto_id:
        return JsonResponse({'error': 'Producto inválido. Usa ID interno, código o código de barras.'}, status=400)

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    rows = list(
        VentaDetalle.objects.filter(venta__in=ventas_qs, producto_id=producto_id)
        .annotate(month=TruncMonth('venta__fecha'))
        .values('month')
        .annotate(units=Sum('cantidad'), avg_price=Avg('precio_unitario'))
        .order_by('month')
    )
    series = [
        {
            'month': (r.get('month').date().isoformat() if r.get('month') else None),
            'units': float(r.get('units') or 0),
            'avg_price': float(r.get('avg_price') or 0),
        }
        for r in rows
        if (r.get('units') or 0) and (r.get('avg_price') or 0)
    ]

    if len(series) < 2:
        return JsonResponse({'producto_id': producto_id, 'available': False, 'message': 'No hay suficientes meses con ventas para estimar elasticidad.', 'series': series})

    a = series[-2]
    b = series[-1]
    if a['units'] <= 0 or a['avg_price'] <= 0:
        return JsonResponse({'producto_id': producto_id, 'available': False, 'message': 'Datos insuficientes para el cálculo.', 'series': series})

    pct_units = ((b['units'] - a['units']) / a['units']) if a['units'] else 0.0
    pct_price = ((b['avg_price'] - a['avg_price']) / a['avg_price']) if a['avg_price'] else 0.0
    elasticity = (pct_units / pct_price) if pct_price else None

    return JsonResponse({
        'producto_id': producto_id,
        'available': elasticity is not None,
        'month_prev': a['month'],
        'month_curr': b['month'],
        'prev': a,
        'curr': b,
        'pct_units': float(pct_units),
        'pct_price': float(pct_price),
        'elasticity': (float(elasticity) if elasticity is not None else None),
        'series': series,
    })


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_new_products_conversion_api(request):
    """New products success rate based on Product.fecha_ingreso_producto and units sold vs baseline."""
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    if not _require_applied_filters(start_dt, end_dt):
        return JsonResponse({'error': 'Debe aplicar filtros (start_date y end_date).'}, status=400)

    end_date = timezone.localdate(end_dt)
    cutoff = end_date - datetime.timedelta(days=90)

    prod_qs = Product.objects.filter(fecha_ingreso_producto__gte=cutoff, fecha_ingreso_producto__lte=end_date)
    if sucursal_id:
        prod_qs = prod_qs.filter(sucursal_id=sucursal_id)

    new_products = list(prod_qs.order_by('fecha_ingreso_producto')[:5000])
    if not new_products:
        return JsonResponse({'available': True, 'cutoff': cutoff.isoformat(), 'new_products_total': 0, 'baseline_units_per_product': 0, 'success': 0, 'fail': 0, 'winners': [], 'losers': []})

    ventas_qs = _ventas_filtered_qs(start_dt, end_dt, sucursal_id=sucursal_id)
    vd = VentaDetalle.objects.filter(venta__in=ventas_qs)

    base = vd.aggregate(total_units=Sum('cantidad'), distinct_products=Count('producto_id', distinct=True))
    total_units = float(base.get('total_units') or 0)
    distinct_products = int(base.get('distinct_products') or 0)
    baseline = (total_units / float(distinct_products)) if distinct_products else 0.0

    units_rows = list(
        vd.filter(producto_id__in=[p.id for p in new_products])
        .values('producto_id', 'producto__nombre')
        .annotate(units=Sum('cantidad'))
    )
    units_by_id = {int(r['producto_id']): float(r.get('units') or 0) for r in units_rows}

    scored = []
    for p in new_products:
        units = float(units_by_id.get(p.id) or 0)
        scored.append({
            'producto_id': int(p.id),
            'producto_nombre': p.nombre or str(p.producto_id) or f'Producto #{p.id}',
            'fecha_ingreso_producto': p.fecha_ingreso_producto.isoformat() if p.fecha_ingreso_producto else None,
            'units': units,
            'success': bool(units >= baseline) if baseline > 0 else bool(units > 0),
        })

    winners = sorted([x for x in scored if x['success']], key=lambda x: x['units'], reverse=True)
    losers = sorted([x for x in scored if not x['success']], key=lambda x: x['units'])

    return JsonResponse({
        'available': True,
        'cutoff': cutoff.isoformat(),
        'new_products_total': int(len(scored)),
        'baseline_units_per_product': float(baseline),
        'success': int(len(winners)),
        'fail': int(len(losers)),
        'winners': winners[:10],
        'losers': losers[:10],
    })


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def market_export_preview(request):
    """Printable (2-page) export preview. Intended for Print/Save to PDF."""
    start_dt, end_dt, sucursal_id = _parse_market_filters(request)
    empleado_id_raw = (request.GET.get('empleado_id') or '').strip()
    empleado_id = int(empleado_id_raw) if empleado_id_raw.isdigit() else None

    if not _require_applied_filters(start_dt, end_dt):
        return HttpResponse('Debe indicar start_date y end_date.', status=400)

    ventas_qs = _ventas_filtered_qs_with_employee(start_dt, end_dt, sucursal_id=sucursal_id, empleado_id=empleado_id)
    core = _aggregate_sales_core(ventas_qs)
    sku_breadth_avg = _avg_distinct_skus_per_ticket(ventas_qs)

    payment = list(
        ventas_qs.values('forma_pago')
        .annotate(total_monto=Sum('total'), tickets=Count('id'))
        .order_by('-total_monto')
    )

    vd = VentaDetalle.objects.filter(venta__in=ventas_qs).select_related('producto')
    top_by_amount = list(
        vd.values('producto_id', 'producto__nombre')
        .annotate(monto_total=Sum(ExpressionWrapper(F('precio_unitario') * F('cantidad'), output_field=DecimalField(max_digits=18, decimal_places=2))))
        .order_by('-monto_total')[:8]
    )
    top_by_qty = list(
        vd.values('producto_id', 'producto__nombre')
        .annotate(cantidad=Sum('cantidad'))
        .order_by('-cantidad')[:8]
    )

    monthly = list(
        ventas_qs.annotate(month=TruncMonth('fecha'))
        .values('month')
        .annotate(total_monto=Sum('total'), tickets=Count('id'))
        .order_by('month')
    )

    annotated = ventas_qs.annotate(day=ExtractDay('fecha'))
    payday = annotated.filter(Q(day__gte=1, day__lte=5) | Q(day__gte=15, day__lte=20)).aggregate(total_monto=Sum('total'), tickets=Count('id'))
    rest = annotated.exclude(Q(day__gte=1, day__lte=5) | Q(day__gte=15, day__lte=20)).aggregate(total_monto=Sum('total'), tickets=Count('id'))
    paycycle = {
        'paydays': {'total_monto': float(payday.get('total_monto') or 0), 'tickets': int(payday.get('tickets') or 0)},
        'rest': {'total_monto': float(rest.get('total_monto') or 0), 'tickets': int(rest.get('tickets') or 0)},
    }

    context = {
        'filters': {
            'start_date': start_dt.date().isoformat(),
            'end_date': end_dt.date().isoformat(),
            'sucursal_id': sucursal_id,
            'empleado_id': empleado_id,
        },
        'kpis': {
            'venta_total_neta': float(core['total_monto']),
            'utilidad_bruta': float(core['gross_profit']),
            'margen_pct': float(core['gross_margin_pct']),
            'ticket_promedio': float(core['avg_ticket']),
            'articulos_por_ticket': float(core['avg_items_per_ticket']),
            'sku_breadth_avg': float(sku_breadth_avg),
            'descuentos_estimado': float(core['total_discount_estimated']),
        },
        'payment_methods': payment,
        'top_by_amount': top_by_amount,
        'top_by_qty': top_by_qty,
        'monthly': monthly,
        'paycycle': paycycle,
        'sucursal': Sucursal.objects.filter(id=sucursal_id).first() if sucursal_id else None,
        'empleado': User.objects.filter(id=empleado_id).first() if empleado_id else None,
    }
    return render(request, 'reports/market_export.html', context)


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def chat_conversations(request):
    """List and create conversations for the current user."""
    if request.method == 'GET':
        convs = Conversation.objects.filter(user=request.user).order_by('-updated_at')[:50]
        data = [
            {'id': c.id, 'title': c.title or f'Conversation {c.id}', 'updated_at': c.updated_at.isoformat()}
            for c in convs
        ]
        return JsonResponse({'conversations': data})
    if request.method == 'POST':
        try:
            body = _json.loads(request.body.decode('utf-8') or '{}')
        except Exception:
            body = {}
        title = body.get('title') or ''
        conv = Conversation.objects.create(user=request.user, title=title)
        return JsonResponse({'id': conv.id, 'title': conv.title})
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def chat_messages(request, conv_id):
    """Get or post messages for a conversation. POST triggers inference and stores assistant reply."""
    conv = get_object_or_404(Conversation, id=conv_id, user=request.user)
    if request.method == 'GET':
        msgs = conv.messages.all().order_by('created_at')
        out = [
            {'id': m.id, 'sender': m.sender, 'text': m.text, 'created_at': m.created_at.isoformat()}
            for m in msgs
        ]
        return JsonResponse({'messages': out})
    if request.method == 'POST':
        try:
            body = _json.loads(request.body.decode('utf-8') or '{}')
        except Exception:
            body = {}
        text = body.get('text', '').strip()
        filters = body.get('filters', {}) or {}
        if not text:
            return JsonResponse({'error': 'Empty text'}, status=400)
        # Save user message
        Message.objects.create(conversation=conv, sender='user', text=text)
        # Build context: reuse compute_analytics with filters
        from .analytics import compute_analytics
        # parse filter dates
        fecha_inicio = None
        fecha_fin = None
        try:
            if filters.get('fecha_inicio'):
                fecha_inicio = timezone.make_aware(datetime.datetime.strptime(filters.get('fecha_inicio'), '%Y-%m-%d'))
            else:
                fecha_inicio = timezone.now() - datetime.timedelta(days=30)
        except Exception:
            fecha_inicio = timezone.now() - datetime.timedelta(days=30)
        try:
            if filters.get('fecha_fin'):
                fecha_fin = timezone.make_aware(datetime.datetime.strptime(filters.get('fecha_fin'), '%Y-%m-%d')) + datetime.timedelta(days=1, seconds=-1)
            else:
                fecha_fin = timezone.now()
        except Exception:
            fecha_fin = timezone.now()
        cajero = filters.get('cajero', 'todos')
        sucursal = filters.get('sucursal', 'todos')
        analytics = compute_analytics(fecha_inicio, fecha_fin, cajero, sucursal)
        # Prepare context summary (small)
        context_summary = {
            'params': {'fecha_inicio': filters.get('fecha_inicio'), 'fecha_fin': filters.get('fecha_fin'), 'cajero': cajero, 'sucursal': sucursal},
            'kpis': {
                'ingreso_total_clp': format_clp(analytics.get('ingreso_total') or 0),
                'ganancia_neta_clp': format_clp(analytics.get('ganancia_neta') or 0),
                'num_transacciones': analytics.get('num_transacciones') or 0,
            },
            'top_products': [
                {'producto': r.get('producto'), 'cantidad': r.get('cantidad')}
                for r in analytics.get('rentabilidad_productos')[:10]
            ]
            if analytics.get('rentabilidad_productos')
            else [],
        }
        # Call inference (public model if configured, else local fallback)
        reply_text = _call_inference(text, context_summary)
        # Save assistant reply
        Message.objects.create(conversation=conv, sender='assistant', text=reply_text, metadata={'context_used': True})
        return JsonResponse({'reply': reply_text})
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@ensure_csrf_cookie
@user_passes_test(_is_admin, login_url='cashier_dashboard')
@login_required(login_url='login')
def report_dashboard(request):
    """Pantalla principal para la generación de reportes"""
    from MOVOS.react_ui import react_ui_enabled, vite_dev_server_url_if_up
    if react_ui_enabled():
        ctx = {}
        if getattr(settings, 'DEBUG', False):
            vite_url = vite_dev_server_url_if_up()
            if vite_url:
                ctx['vite_dev_server_url'] = vite_url
        return render(request, 'reports/react_report_dashboard.html', ctx)
    # Proveer listas para filtros (sucursales y empleados)
    sucursales = Sucursal.objects.all()
    empleados = User.objects.filter(is_active=True)
    return render(request, 'reports/report_dashboard.html', {
        'sucursales': sucursales,
        'empleados': empleados,
    })


@user_passes_test(_is_admin, login_url='cashier_dashboard')
@login_required(login_url='login')
def compute_analytics_api(request):
    """API JSON que recibe filtros (fecha_inicio, fecha_fin, cajero, sucursal)
    y devuelve los KPIs computados por `compute_analytics`.
    Método: POST con JSON.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'method_not_allowed'}, status=405)
    try:
        payload = _json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        payload = {}
    # Parsear fechas: esperar 'YYYY-MM-DD'
    fecha_inicio = payload.get('fecha_inicio')
    fecha_fin = payload.get('fecha_fin')
    cajero = payload.get('cajero', 'todos')
    sucursal = payload.get('sucursal', 'todos')
    try:
        if fecha_inicio:
            fi_dt = timezone.make_aware(datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d'))
        else:
            fi_dt = timezone.now() - datetime.timedelta(days=30)
    except Exception:
        fi_dt = timezone.now() - datetime.timedelta(days=30)
    try:
        if fecha_fin:
            ff_dt = timezone.make_aware(datetime.datetime.strptime(fecha_fin, '%Y-%m-%d')) + datetime.timedelta(days=1, seconds=-1)
        else:
            ff_dt = timezone.now()
    except Exception:
        ff_dt = timezone.now()

    # Calcular analytics
    from .analytics import compute_analytics
    analytics = compute_analytics(fi_dt, ff_dt, cajero_filter=cajero, sucursal_filter=sucursal)

    # Serializar Decimals/dates a tipos JSON-compatibles
    # Regla: montos -> int (pesos CLP), porcentajes/ratios -> float
    def _convert(obj, key=None):
        if isinstance(obj, Decimal):
            return float(obj) if _is_pct_key(key) else _clp_int(obj)
        if isinstance(obj, (list, tuple)):
            return [_convert(i) for i in obj]
        if isinstance(obj, dict):
            return {k: _convert(v, k) for k, v in obj.items()}
        if isinstance(obj, (datetime.date, datetime.datetime)):
            return obj.isoformat()
        return obj

    out = _convert(analytics)
    return JsonResponse({'success': True, 'data': out})


@ensure_csrf_cookie
@login_required(login_url='login')
def refresh_csrf(request):
    """Endpoint simple para re-sincronizar la cookie CSRF en el cliente.

    Uso: GET /reports/refresh-csrf/ -> fuerza que Django emita la cookie `csrftoken`
    y devuelve el token en JSON para clientes que quieran leerlo.
    Requiere sesión autenticada para evitar emisión de tokens a usuarios anónimos.
    """
    try:
        from django.middleware.csrf import get_token
        token = get_token(request)
        logger.info("refresh_csrf called by user=%s (id=%s) - token_emitted=%s", getattr(request.user, 'username', None), getattr(request.user, 'id', None), bool(token))
        return JsonResponse({'success': True, 'csrftoken': token})
    except Exception as e:
        logger.exception("refresh_csrf error for user=%s: %s", getattr(request.user, 'username', None), str(e))
        return JsonResponse({'success': False, 'error': 'could_not_get_token'}, status=500)

@user_passes_test(_is_admin, login_url='cashier_dashboard')
@login_required(login_url='login')
def sales_dashboard(request):
    """Pantalla principal de Gestión de Ventas con opciones"""
    return render(request, 'reports/sales_dashboard.html')

@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def sales_history(request):
    """Historial de Ventas"""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    empleado_id = request.GET.get('empleado')
    sucursal_id = request.GET.get('sucursal')
    numero_transaccion = request.GET.get('numero_transaccion')
    forma_pago = request.GET.get('forma_pago', '')
    page = request.GET.get('page', 1)

    per_page_options = [5, 10, 15, 20]
    per_page = request.GET.get('per_page', 10)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10
    if per_page not in per_page_options:
        per_page = 10

    ventas = Venta.objects.all().order_by('-fecha')

    if fecha_inicio:
        try:
            fecha_inicio_obj = timezone.make_aware(datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d'))
            ventas = ventas.filter(fecha__gte=fecha_inicio_obj)
        except ValueError:
            ventas = Venta.objects.none()
    if fecha_fin:
        try:
            fecha_fin_obj = timezone.make_aware(datetime.datetime.strptime(fecha_fin, '%Y-%m-%d'))
            fecha_fin_obj += datetime.timedelta(days=1, seconds=-1)
            ventas = ventas.filter(fecha__lte=fecha_fin_obj)
        except ValueError:
            ventas = Venta.objects.none()
    if empleado_id:
        try:
            empleado_id = int(empleado_id)
            ventas = ventas.filter(empleado_id=empleado_id)
        except ValueError:
            ventas = Venta.objects.none()
    if sucursal_id:
        try:
            sucursal_id_int = int(sucursal_id)
            ventas = ventas.filter(sucursal_id=sucursal_id_int)
        except ValueError:
            ventas = Venta.objects.none()
    if numero_transaccion:
        numero_transaccion = (numero_transaccion or '').strip()
        if numero_transaccion:
            ventas = ventas.filter(numero_transaccion__icontains=numero_transaccion)
    if forma_pago:
        ventas = ventas.filter(forma_pago=forma_pago)

    # Totales acumulados sobre el queryset completo (antes de paginar)
    _zero = Decimal('0')
    totals_agg = ventas.aggregate(
        total_sum=Sum('total'),
        count=Count('id'),
        ef=Sum(Case(When(forma_pago='efectivo', then=F('total')), default=Value(0), output_field=DecimalField())),
        db=Sum(Case(When(forma_pago='debito', then=F('total')), default=Value(0), output_field=DecimalField())),
        cr=Sum(Case(When(forma_pago='credito', then=F('total')), default=Value(0), output_field=DecimalField())),
        tr=Sum(Case(When(forma_pago='transferencia', then=F('total')), default=Value(0), output_field=DecimalField())),
        mx=Sum(Case(When(forma_pago='mixto', then=F('total')), default=Value(0), output_field=DecimalField())),
    )
    totals = {
        'count': totals_agg['count'] or 0,
        'total_sum': "$" + format_clp(totals_agg['total_sum'] or _zero),
        'efectivo': "$" + format_clp(totals_agg['ef'] or _zero),
        'debito': "$" + format_clp(totals_agg['db'] or _zero),
        'credito': "$" + format_clp(totals_agg['cr'] or _zero),
        'transferencia': "$" + format_clp(totals_agg['tr'] or _zero),
        'mixto': "$" + format_clp(totals_agg['mx'] or _zero),
        'hay_mixto': bool(totals_agg['mx'] and totals_agg['mx'] > 0),
    }

    # Anotar devoluciones para badge
    ventas = ventas.annotate(num_devoluciones=Count('devoluciones'))

    paginator = Paginator(ventas, per_page)
    sales_page = paginator.get_page(page)
    for sale in sales_page:
        sale.display_total = "$" + format_clp(sale.total)
        sale.has_devolucion = sale.num_devoluciones > 0

    empleados = User.objects.all()
    sucursales = Sucursal.objects.all()

    if not numero_transaccion and not fecha_inicio and not fecha_fin:
        today = timezone.localdate()
        tomorrow = today + datetime.timedelta(days=1)
        fecha_inicio = today.isoformat()
        fecha_fin = tomorrow.isoformat()

    return render(request, 'reports/sales_history.html', {
        'sales': sales_page,
        'per_page': per_page,
        'per_page_options': per_page_options,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'empleado_id': empleado_id if empleado_id else None,
        'empleados': empleados,
        'sucursales': sucursales,
        'sucursal_id': sucursal_id if sucursal_id else None,
        'numero_transaccion': numero_transaccion,
        'forma_pago': forma_pago,
        'totals': totals,
    })


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def export_sales_history_excel(request):
    """Exporta el historial de ventas filtrado a Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    empleado_id = request.GET.get('empleado')
    sucursal_id = request.GET.get('sucursal')
    numero_transaccion = (request.GET.get('numero_transaccion') or '').strip()
    forma_pago = request.GET.get('forma_pago', '')

    ventas = Venta.objects.select_related('empleado', 'sucursal').order_by('-fecha')

    if fecha_inicio:
        try:
            ventas = ventas.filter(fecha__gte=timezone.make_aware(datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d')))
        except ValueError:
            ventas = Venta.objects.none()
    if fecha_fin:
        try:
            dt = timezone.make_aware(datetime.datetime.strptime(fecha_fin, '%Y-%m-%d'))
            ventas = ventas.filter(fecha__lte=dt + datetime.timedelta(days=1, seconds=-1))
        except ValueError:
            ventas = Venta.objects.none()
    if empleado_id:
        try:
            ventas = ventas.filter(empleado_id=int(empleado_id))
        except ValueError:
            ventas = Venta.objects.none()
    if sucursal_id:
        try:
            ventas = ventas.filter(sucursal_id=int(sucursal_id))
        except ValueError:
            ventas = Venta.objects.none()
    if numero_transaccion:
        ventas = ventas.filter(numero_transaccion__icontains=numero_transaccion)
    if forma_pago:
        ventas = ventas.filter(forma_pago=forma_pago)

    ventas = ventas.annotate(num_devoluciones=Count('devoluciones'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Historial de Ventas'

    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='0F172A')
    center = Alignment(horizontal='center')

    headers = ['ID', 'Fecha', 'Empleado', 'Sucursal', 'Forma de Pago', 'Nº Transacción', 'Total (CLP)', 'Con Devolución']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center

    LABELS = {
        'efectivo': 'Efectivo', 'debito': 'Tarjeta de Débito',
        'credito': 'Tarjeta de Crédito', 'transferencia': 'Transferencia',
        'nota_credito': 'Nota de crédito', 'mixto': 'Mixto',
    }

    for v in ventas:
        ws.append([
            v.id,
            v.fecha.strftime('%d/%m/%Y %H:%M') if v.fecha else '',
            v.empleado.username if v.empleado else '',
            v.sucursal.nombre if v.sucursal else '',
            LABELS.get(v.forma_pago, v.forma_pago),
            v.numero_transaccion or '',
            int(v.total or 0),
            'Sí' if v.num_devoluciones > 0 else 'No',
        ])

    for i, w in enumerate([8, 18, 16, 16, 20, 22, 14, 15], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f'ventas_{fecha_inicio or "todas"}_{fecha_fin or "todas"}.xlsx'
    resp = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def returns_history(request):
    """Historial de Devoluciones"""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    empleado_id = request.GET.get('empleado')
    sucursal_id = request.GET.get('sucursal')
    numero_transaccion = request.GET.get('numero_transaccion')
    venta_id = request.GET.get('venta_id')
    page = request.GET.get('page', 1)

    per_page_options = [5, 10, 15, 20]
    per_page = request.GET.get('per_page', 10)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10
    if per_page not in per_page_options:
        per_page = 10

    devoluciones = Devolucion.objects.select_related('empleado', 'sucursal', 'venta_original', 'nota_credito').order_by('-fecha')

    if fecha_inicio:
        try:
            fecha_inicio_obj = timezone.make_aware(datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d'))
            devoluciones = devoluciones.filter(fecha__gte=fecha_inicio_obj)
        except ValueError:
            devoluciones = Devolucion.objects.none()
    if fecha_fin:
        try:
            fecha_fin_obj = timezone.make_aware(datetime.datetime.strptime(fecha_fin, '%Y-%m-%d'))
            fecha_fin_obj += datetime.timedelta(days=1, seconds=-1)
            devoluciones = devoluciones.filter(fecha__lte=fecha_fin_obj)
        except ValueError:
            devoluciones = Devolucion.objects.none()
    if empleado_id:
        try:
            empleado_id = int(empleado_id)
            devoluciones = devoluciones.filter(empleado_id=empleado_id)
        except ValueError:
            devoluciones = Devolucion.objects.none()

    if sucursal_id:
        try:
            sucursal_id_int = int(sucursal_id)
            devoluciones = devoluciones.filter(sucursal_id=sucursal_id_int)
        except ValueError:
            devoluciones = Devolucion.objects.none()

    if numero_transaccion:
        numero_transaccion = (numero_transaccion or '').strip()
        if numero_transaccion:
            devoluciones = devoluciones.filter(numero_transaccion__icontains=numero_transaccion)

    if venta_id:
        try:
            venta_id_int = int(venta_id)
            devoluciones = devoluciones.filter(venta_original_id=venta_id_int)
        except ValueError:
            devoluciones = Devolucion.objects.none()

    paginator = Paginator(devoluciones, per_page)
    returns_page = paginator.get_page(page)
    for d in returns_page:
        d.display_total = "$" + format_clp(d.total)
        d.has_nota_credito = bool(d.nota_credito)

    empleados = User.objects.all()
    sucursales = Sucursal.objects.all()

    if not numero_transaccion and not fecha_inicio and not fecha_fin:
        today = timezone.localdate()
        tomorrow = today + datetime.timedelta(days=1)
        fecha_inicio = today.isoformat()
        fecha_fin = tomorrow.isoformat()

    return render(request, 'reports/returns_history.html', {
        'returns': returns_page,
        'per_page': per_page,
        'per_page_options': per_page_options,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'empleado_id': empleado_id if empleado_id else None,
        'empleados': empleados,
        'sucursales': sucursales,
        'sucursal_id': sucursal_id if sucursal_id else None,
        'numero_transaccion': numero_transaccion,
        'venta_id': venta_id,
    })

@ensure_csrf_cookie
@user_passes_test(_is_admin, login_url='cashier_dashboard')
@login_required(login_url='login')
def cash_history(request):
    """Historial de Caja con filtros por ID, cajero y rango de fechas."""
    id_caja_filtro = request.GET.get('id_caja')
    cajero_filtro = request.GET.get('cajero')
    sucursal_filtro = request.GET.get('sucursal')
    fecha_inicio_filtro = request.GET.get('fecha_inicio')
    fecha_fin_filtro = request.GET.get('fecha_fin')
    page = request.GET.get('page', 1)
    per_page_options = [5, 10, 15, 20]
    per_page = request.GET.get('per_page', 10)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10
    if per_page not in per_page_options:
        per_page = 10

    # Se ordena por fecha de apertura (campo "apertura")
    cajas = AperturaCierreCaja.objects.all().order_by('-apertura')
    if id_caja_filtro:
        try:
            cajas = cajas.filter(id=int(id_caja_filtro))
        except ValueError:
            cajas = AperturaCierreCaja.objects.none()
    if cajero_filtro:
        # Filtramos por id de vendedor (select)
        try:
            cajas = cajas.filter(vendedor_id=int(cajero_filtro))
        except Exception:
            # fallback: support username partial match
            boxes = cajas.filter(vendedor__username__icontains=cajero_filtro)
            cajas = boxes
    if sucursal_filtro:
        try:
            cajas = cajas.filter(sucursal_id=int(sucursal_filtro))
        except Exception:
            pass
    if fecha_inicio_filtro:
        try:
            fecha_inicio_obj = parse_date(fecha_inicio_filtro)
            if fecha_inicio_obj:
                cajas = cajas.filter(apertura__gte=fecha_inicio_obj)
        except ValueError:
            cajas = AperturaCierreCaja.objects.none()
    if fecha_fin_filtro:
        try:
            fecha_fin_obj = parse_date(fecha_fin_filtro)
            if fecha_fin_obj:
                cajas = cajas.filter(apertura__lte=fecha_fin_obj)
        except ValueError:
            cajas = AperturaCierreCaja.objects.none()

    paginator = Paginator(cajas, per_page)
    cash_page = paginator.get_page(page)
    # Para cada caja se recalcula el total de ventas (consultando las ventas registradas en el período)
    for caja in cash_page:
        ventas_total = Venta.objects.filter(
            empleado=caja.vendedor,
            fecha__gte=caja.apertura,
            fecha__lte=caja.cierre if caja.cierre else timezone.now()
        ).aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        caja.formatted_ventas_totales = "$" + format_clp(ventas_total)
        # Calcular ventas en efectivo para este periodo y el efectivo final esperado
        ventas_efectivo = Venta.objects.filter(
            empleado=caja.vendedor,
            fecha__gte=caja.apertura,
            fecha__lte=caja.cierre if caja.cierre else timezone.now(),
            forma_pago='efectivo'
        ).aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        efectivo_esperado = (caja.efectivo_inicial or Decimal('0')) + (ventas_efectivo or Decimal('0'))
        # Normalizar a pesos CLP enteros para evitar decimales .00
        efectivo_esperado = efectivo_esperado.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        # Exponer ambos para la plantilla: raw (punto decimal) y formateado
        caja.efectivo_esperado_raw = str(efectivo_esperado)
        caja.formatted_efectivo_esperado = "$" + format_clp(efectivo_esperado)

    # Prepare defaults: start = today, end = tomorrow so the range includes all of today
    today = timezone.localdate()
    tomorrow = today + datetime.timedelta(days=1)
    context = {
        'cajas': cash_page,
        'per_page': per_page,
        'per_page_options': per_page_options,
        'id_caja_filtro': id_caja_filtro,
        'cajero_filtro': cajero_filtro,
        'sucursal_filtro': sucursal_filtro,
        'fecha_inicio_filtro': fecha_inicio_filtro or today.isoformat(),
        'fecha_fin_filtro': fecha_fin_filtro or tomorrow.isoformat(),
        'usuarios': User.objects.all(),
        'sucursales': Sucursal.objects.all(),
    }
    # Calcular la suma total de ventas de todas las cajas que actualmente están abiertas
    try:
        total_open = Decimal('0.00')
        open_cajas_qs = cajas.filter(estado='abierta')
        for oc in open_cajas_qs:
            ventas_total = Venta.objects.filter(
                empleado=oc.vendedor,
                fecha__gte=oc.apertura,
                fecha__lte=timezone.now()
            ).aggregate(total=Sum('total'))['total'] or Decimal('0.00')
            total_open += ventas_total
        context['open_cajas_ventas_total'] = "$" + format_clp(total_open)
    except Exception:
        context['open_cajas_ventas_total'] = "$0"
    return render(request, 'reports/cash_history.html', context)

@user_passes_test(_is_admin, login_url='cashier_dashboard')
@login_required(login_url='login')
def sales_report(request, sale_id):
    """
    Reporte detallado de una venta. Verifica que la venta exista y que tenga detalles.
    Se calculan los valores formateados y se pasan al contexto sin asignarlos a la instancia.
    """
    try:
        venta = get_object_or_404(Venta, id=sale_id)
        detalles = venta.detalles.all()
        logger.info("Venta %s encontrada con %d detalle(s).", sale_id, detalles.count())

        formatted_total = "$" + format_clp(venta.total or 0)
        formatted_cliente_paga = "$" + format_clp(venta.cliente_paga or 0)
        formatted_vuelto_entregado = "$" + format_clp(venta.vuelto_entregado or 0)

        monto_nc_raw = getattr(venta, 'monto_nota_credito', None) or Decimal('0')
        monto_nc = to_clp_pesos(monto_nc_raw)
        total_a_pagar = to_clp_pesos((venta.total or Decimal('0')) - monto_nc_raw)
        formatted_nota_credito = ("$" + format_clp(monto_nc)) if monto_nc and Decimal(str(monto_nc)) > 0 else None
        formatted_total_a_pagar = ("$" + format_clp(total_a_pagar)) if formatted_nota_credito else None

        pagos = list(venta.pagos.all().order_by('id'))
        payment_breakdown = []
        if pagos:
            for p in pagos:
                monto = to_clp_pesos(p.monto or Decimal('0'))
                payment_breakdown.append({
                    'metodo': p.metodo,
                    'metodo_display': p.get_metodo_display() if hasattr(p, 'get_metodo_display') else p.metodo,
                    'formatted_monto': "$" + format_clp(monto),
                    'numero_transaccion': (p.numero_transaccion or '').strip() or None,
                    'banco': (p.banco or '').strip() or None,
                    'nota_credito_codigo': (getattr(p.nota_credito, 'codigo', None) if p.metodo == 'nota_credito' else None),
                })
        else:
            total = to_clp_pesos(venta.total or Decimal('0'))
            monto_nc_int = to_clp_pesos(getattr(venta, 'monto_nota_credito', None) or Decimal('0'))
            restante = to_clp_pesos((venta.total or Decimal('0')) - (getattr(venta, 'monto_nota_credito', None) or Decimal('0')))
            if monto_nc_int and Decimal(str(monto_nc_int)) > 0:
                payment_breakdown.append({
                    'metodo': 'nota_credito',
                    'metodo_display': 'Nota de crédito',
                    'formatted_monto': "$" + format_clp(monto_nc_int),
                    'numero_transaccion': None,
                    'banco': None,
                    'nota_credito_codigo': getattr(getattr(venta, 'nota_credito_usada', None), 'codigo', None),
                })
            if (restante and Decimal(str(restante)) > 0) or not payment_breakdown:
                payment_breakdown.append({
                    'metodo': venta.forma_pago,
                    'metodo_display': venta.get_forma_pago_display(),
                    'formatted_monto': "$" + format_clp(restante if payment_breakdown else total),
                    'numero_transaccion': (venta.numero_transaccion or '').strip() or None,
                    'banco': (venta.banco or '').strip() or None,
                    'nota_credito_codigo': None,
                })

        forma_pago_label = 'Mixta' if venta.forma_pago == 'mixto' else venta.get_forma_pago_display()

        # En lugar de asignar a cada detalle, creamos una lista de diccionarios con los datos formateados
        detalles_formateados = []
        for detalle in detalles:
            subtotal = detalle.cantidad * detalle.precio_unitario
            detalles_formateados.append({
                'producto': detalle.producto,
                'cantidad': detalle.cantidad,
                'precio_unitario': detalle.precio_unitario,
                'formatted_subtotal': "$" + format_clp(subtotal)
            })

        context = {
            'venta': venta,
            'detalles': detalles_formateados,
            'formatted_total': formatted_total,
            'formatted_cliente_paga': formatted_cliente_paga,
            'formatted_vuelto_entregado': formatted_vuelto_entregado,
            'sucursal': venta.sucursal,  # Se agrega la sucursal
            'formatted_nota_credito': formatted_nota_credito,
            'formatted_total_a_pagar': formatted_total_a_pagar,
            'forma_pago_label': forma_pago_label,
            'payment_breakdown': payment_breakdown,
            'show_payment_breakdown': (venta.forma_pago == 'mixto' or len(payment_breakdown) > 1),
        }
        return render(request, 'reports/sales_report.html', context)
    except Exception as e:
        logger.error("Error en sales_report (ID:%s): %s", sale_id, str(e))
        return render(request, 'reports/sales_report.html', {
            'error': str(e),
            'venta': None,
            'detalles': None
        })

@user_passes_test(_is_admin, login_url='cashier_dashboard')
@login_required(login_url='login')
def advanced_reports(request):
    """Legacy advanced reports page.

    Important UX rule: do not show analytics until the user applies filters.
    This view is kept for compatibility with existing flows/tests.
    """
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    cajero = request.GET.get('cajero', 'todos')
    sucursal = request.GET.get('sucursal', 'todos')
    top = request.GET.get('top')
    comparativo_inicio = request.GET.get('comparativo_inicio')
    comparativo_fin = request.GET.get('comparativo_fin')

    show_data = bool(fecha_inicio and fecha_fin)

    context = {
        'usuarios': User.objects.all(),
        'sucursales': Sucursal.objects.all(),
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'cajero_actual': cajero,
        'sucursal_actual': sucursal,
        'top': top or '',
        'comparativo_inicio': comparativo_inicio or '',
        'comparativo_fin': comparativo_fin or '',
        'comparativo_custom': bool(comparativo_inicio and comparativo_fin),
        'show_data': show_data,
    }

    if not show_data:
        return render(request, 'reports/advanced_reports_legacy.html', context)

    # Parse date range
    try:
        fi = timezone.make_aware(datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d'))
    except Exception:
        fi = timezone.now() - datetime.timedelta(days=30)
    try:
        ff = timezone.make_aware(datetime.datetime.strptime(fecha_fin, '%Y-%m-%d')) + datetime.timedelta(days=1, seconds=-1)
    except Exception:
        ff = timezone.now()

    from .analytics import compute_analytics
    analytics = compute_analytics(fi, ff, cajero_filter=cajero, sucursal_filter=sucursal)

    # Promedio de ganancia neta (valor crudo para tests/UI)
    rent = analytics.get('rentabilidad_productos') or []
    promedio_ganancia_neta = Decimal('0.00')
    if rent:
        try:
            promedio_ganancia_neta = (
                sum(Decimal(str(r.get('ganancia_neta_total', 0))) for r in rent) / Decimal(str(len(rent)))
            ).quantize(Decimal('0.01'))
        except Exception:
            promedio_ganancia_neta = Decimal('0.00')

    # Top productos más vendidos (tabla simple para UI)
    try:
        top_n = max(1, int(top)) if top else 10
    except Exception:
        top_n = 10
    top_products = (
        VentaDetalle.objects.filter(venta__fecha__gte=fi, venta__fecha__lte=ff)
        .values('producto__nombre')
        .annotate(cantidad=Sum('cantidad'))
        .order_by('-cantidad')
    )
    if cajero and cajero != 'todos':
        try:
            top_products = top_products.filter(venta__empleado_id=int(cajero))
        except Exception:
            pass
    if sucursal and sucursal != 'todos':
        try:
            top_products = top_products.filter(venta__sucursal_id=int(sucursal))
        except Exception:
            pass
    top_products = list(top_products[:top_n])

    context.update({
        'analytics': analytics,
        'kpis': {
            'ingreso_total': analytics.get('ingreso_total') or Decimal('0.00'),
            'ganancia_neta': analytics.get('ganancia_neta') or Decimal('0.00'),
            'num_transacciones': analytics.get('num_transacciones') or 0,
            'margen': analytics.get('margen') or Decimal('0.00'),
        },
        'promedio_ganancia_neta': promedio_ganancia_neta,
        'promedio_ganancia_neta_raw': str(promedio_ganancia_neta),
        'top_selling_products': top_products,
    })
    return render(request, 'reports/advanced_reports_legacy.html', context)


class AdvancedReportsView(View):
    """Advanced BI dashboard for strategic decisions.
    Renders KPIs and datasets needed for charts. Uses optimized ORM queries.
    """
    def get(self, request):
        # Filters
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        sucursal = request.GET.get('sucursal') or 'todos'

        # Parse dates safely
        try:
            if fecha_inicio:
                fi = timezone.make_aware(datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d'))
            else:
                fi = timezone.now() - datetime.timedelta(days=30)
        except Exception:
            fi = timezone.now() - datetime.timedelta(days=30)
        try:
            if fecha_fin:
                ff = timezone.make_aware(datetime.datetime.strptime(fecha_fin, '%Y-%m-%d')) + datetime.timedelta(days=1, seconds=-1)
            else:
                ff = timezone.now()
        except Exception:
            ff = timezone.now()

        venta_qs = Venta.objects.filter(fecha__gte=fi, fecha__lte=ff).select_related('empleado', 'sucursal')
        if sucursal != 'todos':
            try:
                venta_qs = venta_qs.filter(sucursal_id=int(sucursal))
            except Exception:
                pass

        # KPI: total bruto (con IVA) y total sin IVA (asumiendo 19% IVA)
        totals = venta_qs.aggregate(total_bruto=Sum('total'))
        total_bruto = totals.get('total_bruto') or Decimal('0.00')
        total_neto = (total_bruto / Decimal('1.19')).quantize(Decimal('0.01')) if total_bruto else Decimal('0.00')

        # Transactions count and ticket promedio
        num_trans = venta_qs.count()
        ticket_promedio = (total_bruto / num_trans).quantize(Decimal('0.01')) if num_trans else Decimal('0.00')

        # Ganancia real aggregated via DB expression when possible
        detalles = VentaDetalle.objects.filter(venta__in=venta_qs).select_related('producto')
        try:
            profit_agg = detalles.aggregate(ganancia=Sum(F('cantidad') * (F('precio_unitario') - F('producto__precio_compra'))))
            ganancia_total = profit_agg.get('ganancia') or Decimal('0.00')
        except Exception:
            ganancia_total = Decimal('0.00')
            for d in detalles:
                precio_compra = d.producto.precio_compra or Decimal('0.00')
                ganancia_total += (d.precio_unitario - precio_compra) * d.cantidad

        # Margen promedio %: compute per-venta margin and average
        ventas_marg = venta_qs
        margen_sum = Decimal('0.00')
        margen_count = 0
        for v in ventas_marg:
            ingreso_neto = (v.total / Decimal('1.19')) if v.total else Decimal('0.00')
            dets = VentaDetalle.objects.filter(venta_id=v.id).select_related('producto')
            costo = Decimal('0.00')
            for det in dets:
                costo += (det.producto.precio_compra or Decimal('0.00')) * det.cantidad
            gan_neta = ingreso_neto - (costo / Decimal('1.19')) if costo else ingreso_neto
            if ingreso_neto > 0:
                margen_pct = (gan_neta / ingreso_neto) * Decimal('100')
                margen_sum += margen_pct
                margen_count += 1
        margen_promedio_pct = (margen_sum / margen_count).quantize(Decimal('0.01')) if margen_count else Decimal('0.00')

        # Heatmap: day of week vs hour
        heat_qs = venta_qs.annotate(dow=ExtractWeekDay('fecha'), hr=ExtractHour('fecha')).values('dow','hr').annotate(transacciones=Count('id'), ingreso=Sum('total'))
        heatmap = {}
        for r in heat_qs:
            dow = int(r['dow'])
            hr = int(r['hr'])
            heatmap.setdefault(dow, {})[hr] = {'transacciones': r['transacciones'], 'ingreso': float(r['ingreso'] or 0)}

        # Trend: this month vs previous month
        today = timezone.localdate()
        first_of_month = today.replace(day=1)
        prev_month_end = first_of_month - datetime.timedelta(days=1)
        prev_month_start = prev_month_end.replace(day=1)
        this_month_qs = Venta.objects.filter(fecha__gte=first_of_month, fecha__lte=timezone.now())
        prev_month_qs = Venta.objects.filter(fecha__gte=prev_month_start, fecha__lte=prev_month_end)
        this_daily = this_month_qs.annotate(day=TruncDate('fecha')).values('day').annotate(ingreso=Sum('total')).order_by('day')
        prev_daily = prev_month_qs.annotate(day=TruncDate('fecha')).values('day').annotate(ingreso=Sum('total')).order_by('day')
        trend_this = [{'day': r['day'].isoformat(), 'ingreso': float(r['ingreso'] or 0)} for r in this_daily]
        trend_prev = [{'day': r['day'].isoformat(), 'ingreso': float(r['ingreso'] or 0)} for r in prev_daily]

        # Pareto top products by profit
        prod_profit_qs = VentaDetalle.objects.filter(venta__in=venta_qs).values('producto_id','producto__precio_compra','producto__id','producto__precio_venta','producto__nombre').annotate(total_cantidad=Sum('cantidad'), ingreso=Sum(F('cantidad') * F('precio_unitario')))
        prod_list = []
        total_profit_all = Decimal('0.00')
        for p in prod_profit_qs:
            precio_compra = Decimal(str(p.get('producto__precio_compra') or 0))
            ingreso = Decimal(str(p.get('ingreso') or 0))
            cantidad = Decimal(str(p.get('total_cantidad') or 0))
            costo = precio_compra * cantidad
            profit = ingreso - costo
            prod_list.append({'producto_id': p.get('producto__id'), 'nombre': p.get('producto__nombre'), 'profit': float(profit), 'ingreso': float(ingreso), 'cantidad': float(cantidad)})
            total_profit_all += Decimal(str(profit))
        prod_list.sort(key=lambda x: x['profit'], reverse=True)
        pareto = []
        acc = Decimal('0.00')
        for item in prod_list:
            acc += Decimal(str(item['profit']))
            pareto.append(item)
            if total_profit_all > 0 and (acc / total_profit_all) >= Decimal('0.80'):
                break

        # Sellers performance
        sellers = venta_qs.values('empleado__id', 'empleado__username').annotate(
            total_vendido=Sum('total'),
            tickets=Count('id'),
        ).order_by('-total_vendido')
        sellers_list = [
            {
                'id': s['empleado__id'],
                'username': s['empleado__username'],
                'total_vendido': float(s['total_vendido'] or 0),
                'tickets': int(s['tickets'] or 0),
            }
            for s in sellers
        ]

        # Descuadres negativos
        descuadres = AperturaCierreCaja.objects.filter(cierre__gte=fi, cierre__lte=ff).aggregate(sum_descuadre=Sum('descuadre'))
        sum_descuadre = descuadres.get('sum_descuadre') or Decimal('0.00')

        context = {
            'fecha_inicio': fi.date().isoformat(),
            'fecha_fin': ff.date().isoformat(),
            'total_bruto': float(total_bruto),
            'total_neto': float(total_neto),
            'ganancia_total': float(ganancia_total),
            'ticket_promedio': float(ticket_promedio),
            'margen_promedio_pct': float(margen_promedio_pct),
            'heatmap': heatmap,
            'trend_this': trend_this,
            'trend_prev': trend_prev,
            'pareto_products': pareto,
            'sellers': sellers_list,
            'sum_descuadre': float(sum_descuadre),
            'sucursales': Sucursal.objects.all(),
        }

        return render(request, 'reports/advanced_reports_new.html', context)

@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def advanced_reports_data(request):
    """Nuevo endpoint JSON completo usando compute_analytics sin renderizar HTML."""
    from .analytics import compute_analytics
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    cajero_filter = request.GET.get('cajero','todos')
    sucursal_filter = request.GET.get('sucursal','todos')
    try:
        if fecha_inicio_str:
            fecha_inicio = timezone.make_aware(datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d'))
        else:
            fecha_inicio = timezone.now() - datetime.timedelta(days=30)
    except ValueError:
        fecha_inicio = timezone.now() - datetime.timedelta(days=30)
    try:
        if fecha_fin_str:
            fecha_fin = timezone.make_aware(datetime.datetime.strptime(fecha_fin_str, '%Y-%m-%d')) + datetime.timedelta(days=1, seconds=-1)
        else:
            fecha_fin = timezone.now()
    except ValueError:
        fecha_fin = timezone.now()
    comp_inicio_str = request.GET.get('comparativo_inicio')
    comp_fin_str = request.GET.get('comparativo_fin')
    analytics = compute_analytics(fecha_inicio, fecha_fin, cajero_filter, sucursal_filter)
    # Parametro opcional para limitar Top de productos (default 10)
    filtro_top = request.GET.get('top', 10)
    try:
        filtro_top = int(filtro_top)
    except ValueError:
        filtro_top = 10
    comparativo_meta = {
        'comparativo_custom': False,
        'comparativo_inicio': analytics['prev_inicio'].strftime('%Y-%m-%d'),
        'comparativo_fin': analytics['prev_fin'].strftime('%Y-%m-%d'),
        'ingreso_prev': _clp_int(analytics['ingreso_prev']),
        'ganancia_neta_prev': _clp_int(analytics['ganancia_neta_prev']),
        'num_transacciones_prev': analytics['num_transacciones_prev'],
        'margen_prev': float(analytics['margen_prev']),
        'ingreso_delta': _clp_int(analytics['ingreso_delta']),
        'ingreso_pct': float(analytics['ingreso_pct']),
        'ganancia_neta_delta': _clp_int(analytics['ganancia_neta_delta']),
        'ganancia_neta_pct': float(analytics['ganancia_neta_pct']),
        'transacciones_delta': int(analytics['transacciones_delta']),
        'transacciones_pct': float(analytics['transacciones_pct']),
        'margen_delta': float(analytics['margen_delta']),
        'margen_pct': float(analytics['margen_pct']),
        'participacion_ingreso': float((analytics['ingreso_prev']/analytics['ingreso_total']*Decimal('100')).quantize(Decimal('0.01')) if analytics['ingreso_total'] else Decimal('0.00')),
        'participacion_ganancia': float((analytics['ganancia_neta_prev']/analytics['ganancia_neta']*Decimal('100')).quantize(Decimal('0.01')) if analytics['ganancia_neta'] else Decimal('0.00')),
    }
    if comp_inicio_str and comp_fin_str:
        try:
            comp_inicio = timezone.make_aware(datetime.datetime.strptime(comp_inicio_str, '%Y-%m-%d'))
            comp_fin = timezone.make_aware(datetime.datetime.strptime(comp_fin_str, '%Y-%m-%d')) + datetime.timedelta(days=1, seconds=-1)
            if comp_inicio <= comp_fin:
                ventas_prev_custom = Venta.objects.filter(fecha__gte=comp_inicio, fecha__lte=comp_fin)
                if cajero_filter and cajero_filter != 'todos':
                    try: ventas_prev_custom = ventas_prev_custom.filter(empleado_id=int(cajero_filter))
                    except ValueError: pass
                if sucursal_filter and sucursal_filter != 'todos':
                    try: ventas_prev_custom = ventas_prev_custom.filter(sucursal_id=int(sucursal_filter))
                    except ValueError: pass
                ingreso_prev = ventas_prev_custom.aggregate(t=Sum('total'))['t'] or Decimal('0.00')
                detalles_prev = VentaDetalle.objects.filter(venta__in=ventas_prev_custom)
                cmv_prev_val = detalles_prev.aggregate(cmv_prev=Sum(F('cantidad') * F('producto__precio_compra')))['cmv_prev'] or 0
                cmv_prev = Decimal(str(cmv_prev_val))
                ingreso_sin_iva_prev = (ingreso_prev / Decimal('1.19')).quantize(Decimal('0.01')) if ingreso_prev else Decimal('0.00')
                costo_prev_net = (cmv_prev / Decimal('1.19')).quantize(Decimal('0.01')) if cmv_prev else Decimal('0.00')
                ganancia_neta_prev = ingreso_sin_iva_prev - costo_prev_net
                num_transacciones_prev = ventas_prev_custom.count()
                margen_prev = ((ganancia_neta_prev / ingreso_sin_iva_prev) * Decimal('100')).quantize(Decimal('0.01')) if ingreso_sin_iva_prev > 0 else Decimal('0.00')
                def delta_pct(current: Decimal, previous: Decimal):
                    delta = (current - previous)
                    if previous == 0:
                        pct = Decimal('100.00') if current > 0 else Decimal('0.00')
                    else:
                        pct = ((delta / previous) * Decimal('100')).quantize(Decimal('0.01'))
                    return delta, pct
                ingreso_delta, ingreso_pct = delta_pct(analytics['ingreso_total'], ingreso_prev)
                ganancia_neta_delta, ganancia_neta_pct = delta_pct(analytics['ganancia_neta'], ganancia_neta_prev)
                transacciones_delta, transacciones_pct = delta_pct(Decimal(analytics['num_transacciones']), Decimal(num_transacciones_prev))
                margen_delta, margen_pct = delta_pct(analytics['margen'], margen_prev)
                comparativo_meta.update({
                    'comparativo_custom': True,
                    'comparativo_inicio': comp_inicio.strftime('%Y-%m-%d'),
                    'comparativo_fin': comp_fin.strftime('%Y-%m-%d'),
                    'ingreso_prev': _clp_int(ingreso_prev),
                    'ganancia_neta_prev': _clp_int(ganancia_neta_prev),
                    'num_transacciones_prev': num_transacciones_prev,
                    'margen_prev': float(margen_prev),
                    'ingreso_delta': _clp_int(ingreso_delta),
                    'ingreso_pct': float(ingreso_pct),
                    'ganancia_neta_delta': _clp_int(ganancia_neta_delta),
                    'ganancia_neta_pct': float(ganancia_neta_pct),
                    'transacciones_delta': int(transacciones_delta),
                    'transacciones_pct': float(transacciones_pct),
                    'margen_delta': float(margen_delta),
                    'margen_pct': float(margen_pct),
                    'participacion_ingreso': float((ingreso_prev/analytics['ingreso_total']*Decimal('100')).quantize(Decimal('0.01')) if analytics['ingreso_total'] else Decimal('0.00')),
                    'participacion_ganancia': float((ganancia_neta_prev/analytics['ganancia_neta']*Decimal('100')).quantize(Decimal('0.01')) if analytics['ganancia_neta'] else Decimal('0.00')),
                })
        except ValueError:
            pass
    # Formateo monetario liviano en JSON (sin símbolos para facilitar consumo externo)
    def dec_to_number(d, key=None):
        if isinstance(d, Decimal):
            return float(d) if _is_pct_key(key) else _clp_int(d)
        return d
    json_payload = {
        'params': {
            'fecha_inicio': fecha_inicio_str,
            'fecha_fin': fecha_fin_str,
            'cajero': cajero_filter,
            'sucursal': sucursal_filter,
            'top': filtro_top,
        },
        'kpis': {
            'ingreso_total': _clp_int(analytics['ingreso_total']),
            'ingreso_total_clp': _clp_plain(analytics['ingreso_total']),
            'ingreso_total_sin_iva': _clp_int(analytics['ingreso_total_sin_iva']),
            'ingreso_total_sin_iva_clp': _clp_plain(analytics['ingreso_total_sin_iva']),
            'iva_total': _clp_int(analytics['iva_total_calc']),
            'iva_total_clp': _clp_plain(analytics['iva_total_calc']),
            'ganancia_bruta': _clp_int(analytics['ganancia_bruta']),
            'ganancia_bruta_clp': _clp_plain(analytics['ganancia_bruta']),
            'ganancia_neta': _clp_int(analytics['ganancia_neta']),
            'ganancia_neta_clp': _clp_plain(analytics['ganancia_neta']),
            'margen_pct': float(analytics['margen']),
            'costo_total': _clp_int(analytics['costo_total']),
            'costo_total_clp': _clp_plain(analytics['costo_total']),
            'num_transacciones': analytics['num_transacciones'],
            'ticket_promedio': _clp_int(analytics['ticket_promedio']),
            'ticket_promedio_clp': _clp_plain(analytics['ticket_promedio']),
            'unidades_promedio': analytics['unidades_promedio'],
            'best_selling_product': analytics['best_selling_product'],
            'best_selling_quantity': analytics['best_selling_quantity'],
        },
        'comparativo': {
            'prev_inicio': analytics['prev_inicio'].strftime('%Y-%m-%d'),
            'prev_fin': analytics['prev_fin'].strftime('%Y-%m-%d'),
            'ingreso_prev': _clp_int(analytics['ingreso_prev']),
            'ingreso_prev_clp': _clp_plain(analytics['ingreso_prev']),
            'ganancia_neta_prev': _clp_int(analytics['ganancia_neta_prev']),
            'ganancia_neta_prev_clp': _clp_plain(analytics['ganancia_neta_prev']),
            'num_transacciones_prev': analytics['num_transacciones_prev'],
            'margen_prev': float(analytics['margen_prev']),
            'ingreso_delta': _clp_int(analytics['ingreso_delta']),
            'ingreso_delta_clp': _clp_plain(analytics['ingreso_delta']),
            'ingreso_pct': float(analytics['ingreso_pct']),
            'ganancia_neta_delta': _clp_int(analytics['ganancia_neta_delta']),
            'ganancia_neta_delta_clp': _clp_plain(analytics['ganancia_neta_delta']),
            'ganancia_neta_pct': float(analytics['ganancia_neta_pct']),
            'transacciones_delta': int(analytics['transacciones_delta']),
            'transacciones_pct': float(analytics['transacciones_pct']),
            'margen_delta': float(analytics['margen_delta']),
            'margen_pct': float(analytics['margen_pct']),
        },
        'series': {
            'daily_chart': analytics['daily_chart'],
            'branch_comparison': analytics['branch_comparison'],
            'hourly_distribution': analytics['hourly_distribution'],
            'heatmap_matrix': analytics['heatmap_matrix'],
            'wave_labels': analytics['wave_labels'],
            'wave_gains': analytics['wave_gains'],
        },
        'rentabilidad_productos': analytics['rentabilidad_productos'],
        'ranking_cajeros': analytics['ranking_cajeros'],
        'comparativo_meta': comparativo_meta,
        # Top productos más vendidos (filtrado por parámetro 'top')
        'top_selling_products': list(
            VentaDetalle.objects.filter(
                venta__fecha__gte=fecha_inicio,
                venta__fecha__lte=fecha_fin,
                **({} if cajero_filter in (None, '', 'todos') else {'venta__empleado_id': cajero_filter}),
                **({} if sucursal_filter in (None, '', 'todos') else {'venta__sucursal_id': sucursal_filter}),
            ).values('producto__nombre')
             .annotate(total_cantidad=Sum('cantidad'))
             .order_by('-total_cantidad')[:filtro_top]
        ),
    }
    return JsonResponse(json_payload)

@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def export_rentabilidad_csv(request):
    # Reutilizar rango de fechas principal
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    try:
        if fecha_inicio_str:
            fecha_inicio = timezone.make_aware(datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d'))
        else:
            fecha_inicio = timezone.now() - datetime.timedelta(days=30)
    except ValueError:
        fecha_inicio = timezone.now() - datetime.timedelta(days=30)
    try:
        if fecha_fin_str:
            fecha_fin = timezone.make_aware(datetime.datetime.strptime(fecha_fin_str, '%Y-%m-%d')) + datetime.timedelta(days=1, seconds=-1)
        else:
            fecha_fin = timezone.now()
    except ValueError:
        fecha_fin = timezone.now()
    ventas_qs = Venta.objects.filter(fecha__gte=fecha_inicio, fecha__lte=fecha_fin)
    detalles_rango = VentaDetalle.objects.filter(venta__in=ventas_qs)
    rent_map = {}
    for det in detalles_rango.select_related('producto'):
        pid = det.producto_id
        venta_unit = det.precio_unitario or Decimal('0.00')
        compra_unit = det.producto.precio_compra or Decimal('0.00')
        venta_sin_iva = (venta_unit / Decimal('1.19')).quantize(Decimal('0.01')) if venta_unit else Decimal('0.00')
        compra_sin_iva = (compra_unit / Decimal('1.19')).quantize(Decimal('0.01')) if compra_unit else Decimal('0.00')
        ganancia_unit = venta_sin_iva - compra_sin_iva
        entry = rent_map.get(pid)
        if not entry:
            entry = {
                'producto': det.producto.nombre or det.producto.producto_id,
                'cantidad': 0,
                'ingreso': Decimal('0.00'),
                'costo': Decimal('0.00'),
                'ganancia': Decimal('0.00')
            }
        entry['cantidad'] += det.cantidad
        entry['ingreso'] += venta_sin_iva * det.cantidad
        entry['costo'] += compra_sin_iva * det.cantidad
        entry['ganancia'] += ganancia_unit * det.cantidad
        rent_map[pid] = entry
    rows = []
    for _, d in rent_map.items():
        pct = Decimal('0.00')
        if d['ingreso'] > 0:
            pct = (d['ganancia'] / d['ingreso'] * Decimal('100')).quantize(Decimal('0.01'))
        rows.append([d['producto'], d['cantidad'], d['ingreso'], d['costo'], d['ganancia'], pct])
    rows.sort(key=lambda r: r[4], reverse=True)
    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename=rentabilidad_productos.csv'
    # BOM para que Excel detecte UTF-8 correctamente
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['Producto','Cantidad','Ingreso Neto','Costo Neto','Ganancia Neta','% Ganancia'])
    for r in rows:
        writer.writerow([r[0], r[1], r[2], r[3], r[4], r[5]])
    return response

@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def export_ranking_cajeros_csv(request):
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    try:
        if fecha_inicio_str:
            fecha_inicio = timezone.make_aware(datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d'))
        else:
            fecha_inicio = timezone.now() - datetime.timedelta(days=30)
    except ValueError:
        fecha_inicio = timezone.now() - datetime.timedelta(days=30)
    try:
        if fecha_fin_str:
            fecha_fin = timezone.make_aware(datetime.datetime.strptime(fecha_fin_str, '%Y-%m-%d')) + datetime.timedelta(days=1, seconds=-1)
        else:
            fecha_fin = timezone.now()
    except ValueError:
        fecha_fin = timezone.now()
    ventas_qs = Venta.objects.filter(fecha__gte=fecha_inicio, fecha__lte=fecha_fin).select_related('empleado')
    ranking_map = {}
    for v in ventas_qs:
        uid = v.empleado_id
        entry = ranking_map.get(uid)
        if not entry:
            entry = {'usuario': v.empleado.username, 'ventas': 0, 'ingreso': Decimal('0.00')}
        entry['ventas'] += 1
        entry['ingreso'] += v.total or Decimal('0.00')
        ranking_map[uid] = entry
    rows = []
    for _, d in ranking_map.items():
        ticket_prom = (d['ingreso'] / d['ventas']).quantize(Decimal('0.01')) if d['ventas'] else Decimal('0.00')
        rows.append([d['usuario'], d['ventas'], d['ingreso'], ticket_prom])
    rows.sort(key=lambda r: r[2], reverse=True)
    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename=ranking_cajeros.csv'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['Usuario','Ventas','Ingreso Total','Ticket Promedio'])
    for r in rows:
        writer.writerow([r[0], r[1], r[2], r[3]])
    return response


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def export_analytics_csv(request):
    """Exporta un CSV con KPIs principales y top productos para el rango solicitado."""
    from .analytics import compute_analytics
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    cajero_filter = request.GET.get('cajero','todos')
    sucursal_filter = request.GET.get('sucursal','todos')
    try:
        if fecha_inicio_str:
            fecha_inicio = timezone.make_aware(datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d'))
        else:
            fecha_inicio = timezone.now() - datetime.timedelta(days=30)
    except ValueError:
        fecha_inicio = timezone.now() - datetime.timedelta(days=30)
    try:
        if fecha_fin_str:
            fecha_fin = timezone.make_aware(datetime.datetime.strptime(fecha_fin_str, '%Y-%m-%d')) + datetime.timedelta(days=1, seconds=-1)
        else:
            fecha_fin = timezone.now()
    except ValueError:
        fecha_fin = timezone.now()

    analytics = compute_analytics(fecha_inicio, fecha_fin, cajero_filter, sucursal_filter, limit_rentabilidad=100)
    # Top productos
    top_products = VentaDetalle.objects.filter(venta__fecha__gte=fecha_inicio, venta__fecha__lte=fecha_fin)
    if cajero_filter and cajero_filter != 'todos':
        try:
            top_products = top_products.filter(venta__empleado_id=int(cajero_filter))
        except ValueError:
            pass
    if sucursal_filter and sucursal_filter != 'todos':
        try:
            top_products = top_products.filter(venta__sucursal_id=int(sucursal_filter))
        except ValueError:
            pass
    top_list = list(top_products.values('producto__nombre').annotate(total_cantidad=Sum('cantidad')).order_by('-total_cantidad')[:50])

    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename=analytics_kpis.csv'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    # Escribir KPIs principales
    writer.writerow(['KPI','Valor'])
    writer.writerow(['Ingreso Total (CLP)', analytics['ingreso_total']])
    writer.writerow(['Ingreso Neto (sin IVA)', analytics['ingreso_total_sin_iva']])
    writer.writerow(['IVA Calculado', analytics['iva_total_calc']])
    writer.writerow(['Ganancia Bruta', analytics['ganancia_bruta']])
    writer.writerow(['Ganancia Neta', analytics['ganancia_neta']])
    writer.writerow(['Margen (%)', analytics['margen']])
    writer.writerow(['Numero Transacciones', analytics['num_transacciones']])
    writer.writerow(['Ticket Promedio', analytics['ticket_promedio']])
    writer.writerow([])
    writer.writerow(['Top Productos','Cantidad'])
    for p in top_list:
        writer.writerow([p.get('producto__nombre'), p.get('total_cantidad')])
    return response
    return response

@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def export_daily_series_csv(request):
    """Exporta la serie diaria (ingreso y ganancia neta) en CSV."""
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    cajero_filter = request.GET.get('cajero','todos')
    sucursal_filter = request.GET.get('sucursal','todos')
    try:
        if fecha_inicio_str:
            fecha_inicio = timezone.make_aware(datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d'))
        else:
            fecha_inicio = timezone.now() - datetime.timedelta(days=30)
    except ValueError:
        fecha_inicio = timezone.now() - datetime.timedelta(days=30)
    try:
        if fecha_fin_str:
            fecha_fin = timezone.make_aware(datetime.datetime.strptime(fecha_fin_str, '%Y-%m-%d')) + datetime.timedelta(days=1, seconds=-1)
        else:
            fecha_fin = timezone.now()
    except ValueError:
        fecha_fin = timezone.now()
    from .analytics import compute_analytics
    analytics = compute_analytics(fecha_inicio, fecha_fin, cajero_filter, sucursal_filter)
    rows = analytics['daily_chart']
    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename=serie_diaria.csv'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['Dia','Ingreso','GananciaNeta'])
    for r in rows:
        writer.writerow([r['day'], r['ingreso'], r['ganancia_neta']])
    return response

@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def export_branch_comparison_csv(request):
    """Exporta comparación por sucursal (ingreso y ganancia neta) en CSV."""
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    cajero_filter = request.GET.get('cajero','todos')
    sucursal_filter = request.GET.get('sucursal','todos')
    try:
        if fecha_inicio_str:
            fecha_inicio = timezone.make_aware(datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d'))
        else:
            fecha_inicio = timezone.now() - datetime.timedelta(days=30)
    except ValueError:
        fecha_inicio = timezone.now() - datetime.timedelta(days=30)
    try:
        if fecha_fin_str:
            fecha_fin = timezone.make_aware(datetime.datetime.strptime(fecha_fin_str, '%Y-%m-%d')) + datetime.timedelta(days=1, seconds=-1)
        else:
            fecha_fin = timezone.now()
    except ValueError:
        fecha_fin = timezone.now()
    from .analytics import compute_analytics
    analytics = compute_analytics(fecha_inicio, fecha_fin, cajero_filter, sucursal_filter)
    rows = analytics['branch_comparison']
    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename=comparacion_sucursal.csv'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['Sucursal','Ingreso','GananciaNeta'])
    for r in rows:
        writer.writerow([r['sucursal'], r['ingreso'], r['ganancia_neta']])
    return response

@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def caja_report(request, caja_id):
    """
    Vista para mostrar el detalle de una caja cerrada usando el template 'reporte_caja.html'.
    """
    caja = get_object_or_404(AperturaCierreCaja, id=caja_id)
    # Usar los totales persistidos en la caja cuando estén disponibles
    caja.formatted_efectivo_inicial = "$" + format_clp(caja.efectivo_inicial or 0)
    caja.formatted_total_ventas_debito = "$" + format_clp(caja.total_ventas_debito or 0)
    caja.formatted_total_ventas_credito = "$" + format_clp(caja.total_ventas_credito or 0)
    caja.formatted_total_ventas_efectivo = "$" + format_clp(caja.total_ventas_efectivo or 0)
    caja.formatted_vuelto_entregado = "$" + format_clp(caja.vuelto_entregado or 0)
    caja.formatted_efectivo_final = "$" + format_clp(caja.efectivo_final or caja.efectivo_inicial or 0)
    caja.formatted_ventas_totales = "$" + format_clp(caja.ventas_totales or 0)
    # Para transferencias mantenemos el cálculo puntual (no hay campo persistido para transferencias)
    ventas_transferencia = Venta.objects.filter(caja=caja, forma_pago='transferencia').aggregate(total=Sum('total'))['total'] or Decimal('0.00')
    caja.formatted_total_ventas_transferencia = "$" + format_clp(ventas_transferencia)
    
    return render(request, 'reports/reporte_caja.html', {'caja': caja})

@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
@require_http_methods(["POST"])
def limpiar_historial_ventas(request):
    try:
        Venta.objects.all().delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
@require_http_methods(["POST"])
def limpiar_historial_caja(request):
    AperturaCierreCaja.objects.all().delete()
    return JsonResponse({'success': True})


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def export_advanced_pdf(request):
    """Exporta el reporte avanzado completo a PDF usando la misma data y un template html compacto."""
    return HttpResponse('Export a PDF for advanced reports is deprecated.', status=410)


@login_required
@user_passes_test(_is_admin, login_url='cashier_dashboard')
def export_advanced_docx(request):
    """Exporta un resumen del reporte avanzado a DOCX."""
    return HttpResponse('Export a DOCX for advanced reports is deprecated.', status=410)

