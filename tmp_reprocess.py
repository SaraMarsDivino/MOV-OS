from openpyxl import load_workbook
from products.models import Product
from django.db import transaction
from MOVOS.money import parse_clp_pesos
from decimal import Decimal
from datetime import datetime, date
from django.utils.dateparse import parse_date
import re

fp = 'subir.xlsx'
print('Abriendo', fp)
wb = load_workbook(fp, read_only=True, data_only=True)
sh = wb.active
first_row = next(sh.iter_rows(min_row=1, max_row=1, values_only=True), None)
headers = [str(v).strip() if v is not None else '' for v in (first_row or [])]
print('Headers:', headers)

def _norm(h):
    if not h:
        return ''
    h2 = re.sub(r"\s*\(.*\)\s*$", '', str(h)).strip()
    return h2.upper()

header_map = { _norm(headers[i]): i for i in range(len(headers)) if headers[i] }
print('Normalized header keys:', list(header_map.keys()))

minimal = ['NOMBRE','CODIGO 1','PRECIO DE COMPRA','PRECIO DE VENTA']
missing = [h for h in minimal if h not in header_map]
if missing:
    print('Faltan encabezados obligatorios:', missing)
    wb.close()
else:
    existing_map = {p.producto_id: p for p in Product.objects.filter(producto_id__isnull=False)}
    parsed_map = {}
    rows_total = 0
    for row_idx, row_values in enumerate(sh.iter_rows(min_row=2, values_only=True), start=2):
        rows_total += 1
        if not any(v for v in row_values if v is not None and str(v).strip() != ''):
            continue
        def get_val(h):
            idx = header_map.get(h if h is None else h.upper())
            if idx is None:
                idx = header_map.get(re.sub(r"\s*\(.*\)\s*$", '', str(h)).upper())
            if idx is not None and idx < len(row_values):
                return row_values[idx]
            return None
        prodcode = get_val('CODIGO 1')
        if prodcode is None or str(prodcode).strip()=='':
            continue
        prodcode = str(prodcode).strip()
        nombre = str(get_val('NOMBRE')).strip() if get_val('NOMBRE') is not None else ''
        descripcion = str(get_val('DESCRIPCION')).strip() if get_val('DESCRIPCION') is not None else ''
        codigo_alternativo = None
        if 'CODIGO 2' in header_map:
            raw = get_val('CODIGO 2')
            if raw is not None and str(raw).strip() != '':
                codigo_alternativo = str(raw).strip()
        codigo_barras = None
        if 'CODIGO DE BARRAS' in header_map:
            raw = get_val('CODIGO DE BARRAS')
            if raw is not None and str(raw).strip() != '':
                codigo_barras = str(raw).strip()
        if not codigo_barras and codigo_alternativo:
            codigo_barras = codigo_alternativo
        fecha = None
        fecha_raw = get_val('FECHA DE INGRESO')
        if fecha_raw is not None and str(fecha_raw).strip() != '':
            if isinstance(fecha_raw, (datetime, date)):
                fecha = fecha_raw.date() if isinstance(fecha_raw, datetime) else fecha_raw
            else:
                try:
                    fecha = parse_date(str(fecha_raw).split(' ')[0].strip())
                except Exception:
                    fecha = None
        def safe_decimal(v):
            if v is None or (isinstance(v, str) and str(v).strip()==''):
                return Decimal('0')
            try:
                return parse_clp_pesos(v)
            except Exception:
                try:
                    return Decimal(str(v))
                except Exception:
                    return Decimal('0')
        precio_compra = safe_decimal(get_val('PRECIO DE COMPRA'))
        precio_venta = safe_decimal(get_val('PRECIO DE VENTA'))
        parsed_map[prodcode] = {
            'nombre': nombre,
            'descripcion': descripcion or None,
            'codigo_alternativo': codigo_alternativo,
            'codigo_barras': codigo_barras,
            'fecha_ingreso_producto': fecha,
            'precio_compra': precio_compra,
            'precio_venta': precio_venta,
            'permitir_venta_sin_stock': True,
        }
    # build create/update lists
    to_create = []
    to_update = []
    unchanged = []
    for codigo, defaults in parsed_map.items():
        if codigo in existing_map:
            prod = existing_map[codigo]
            changed = False
            for k, v in defaults.items():
                if isinstance(v, str) and v.strip() == '':
                    nv = None
                else:
                    nv = v
                if nv is not None and getattr(prod, k) != nv:
                    setattr(prod, k, nv)
                    changed = True
            if changed:
                to_update.append(prod)
            else:
                unchanged.append(codigo)
        else:
            to_create.append(Product(producto_id=codigo, **defaults))
    created = 0
    updated = 0
    try:
        with transaction.atomic():
            if to_create:
                Product.objects.bulk_create(to_create, batch_size=500)
                created = len(to_create)
            if to_update:
                Product.objects.bulk_update(to_update, ['nombre','descripcion','codigo_alternativo','codigo_barras','fecha_ingreso_producto','precio_compra','precio_venta','permitir_venta_sin_stock'], batch_size=500)
                updated = len(to_update)
    except Exception as e:
        print('Error transacción', e)
    try:
        wb.close()
    except Exception:
        pass
    print('Resumen: filas leídas', rows_total, 'distintos', len(parsed_map), 'creados', created, 'actualizados', updated, 'sin cambios', len(unchanged))
    # show counts
    from products.models import Product as P
    total = P.objects.count()
    with_alt = P.objects.filter(codigo_alternativo__isnull=False).exclude(codigo_alternativo='').count()
    with_bar = P.objects.filter(codigo_barras__isnull=False).exclude(codigo_barras='').count()
    both = P.objects.filter(codigo_alternativo__isnull=False).exclude(codigo_alternativo='').filter(codigo_barras__isnull=False).exclude(codigo_barras='').count()
    print('Total BD', total, 'with_alt', with_alt, 'with_bar', with_bar, 'both', both)
    try:
        p8 = P.objects.get(pk=8)
        print('pk8:', p8.producto_id, p8.nombre, p8.codigo_alternativo, p8.codigo_barras)
    except Exception:
        print('pk8 not found')
